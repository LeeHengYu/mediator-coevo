import csv
import glob
import zipfile
from pathlib import Path

from openpyxl import load_workbook

EXCEL_FILE = Path("/root/output/result.xlsx")
CSV_PATTERN = "/root/output/sheet.csv.*"
COLUMNS = ["H", "I", "J", "K", "L"]
TOLERANCE = 0.02
YELLOW_RGBS = {"00FFF2CC", "FFF2CC", "00FFF2CC"}

EXPECTED_TOP = {"12": [612.5, 628.0, 645.6, 662.3, 679.2], "13": [540.1, 552.8, 565.6, 578.9, 592.5], "14": [488.0, 502.5, 517.1, 531.4, 546.0], "15": [575.4, 589.2, 603.3, 617.8, 632.6], "16": [460.7, 470.5, 480.8, 491.6, 502.8], "17": [520.4, 535.7, 551.0, 566.1, 581.5]}
EXPECTED_MID = {"19": [580.1, 596.2, 612.7, 629.9, 647.0], "20": [553.8, 561.1, 568.1, 574.6, 580.6], "21": [502.4, 511.8, 520.7, 528.8, 536.0], "22": [530.5, 541.7, 553.0, 564.0, 574.6], "23": [470.8, 478.6, 486.1, 493.2, 500.0], "24": [501.4, 519.1, 537.5, 556.3, 575.2]}
EXPECTED_BASE = {"26": [720.0, 739.0, 758.0, 777.0, 796.0], "27": [610.0, 624.0, 638.0, 652.0, 666.0], "28": [540.0, 556.0, 572.0, 588.0, 604.0], "29": [695.0, 711.0, 727.0, 744.0, 761.0], "30": [520.0, 533.0, 546.0, 559.0, 572.0], "31": [640.0, 658.0, 676.0, 694.0, 712.0]}
EXPECTED_NET = {"35": [4.499999999999997, 4.303112313937747, 4.340369393139839, 4.169884169884167, 4.045226130653272], "36": [-2.245901639344251, -1.330128205128216, -0.3918495297805642, 0.6595092024539807, 1.7867867867867833], "37": [-2.6666666666666625, -1.672661870503599, -0.6293706293706334, 0.4421768707483032, 1.6556291390728477], "38": [6.460431654676256, 6.680731364275667, 6.918844566712511, 7.231182795698919, 7.6215505913272015], "39": [-1.9423076923076967, -1.5196998123827434, -0.9706959706959728, -0.28622540250446615, 0.4895104895104915], "40": [2.96875, 2.5227963525835904, 1.9970414201183433, 1.4121037463977044, 0.8848314606741509]}
EXPECTED_STATS = {"42": [-2.6666666666666625, -1.672661870503599, -0.9706959706959728, -0.28622540250446615, 0.4895104895104915], "43": [6.460431654676256, 6.680731364275667, 6.918844566712511, 7.231182795698919, 7.6215505913272015], "44": [0.5132211538461516, 0.5963340737276872, 0.8025959451688895, 1.0358064744258426, 1.7212079629298156], "45": [1.1790509427262739, 1.4973583571304079, 1.8773898750205873, 2.271438563779768, 2.747255766337458], "46": [-2.1700031525851124, -1.4723069105691116, -0.5699903544731161, 0.49650995367472256, 1.077530880273825], "47": [4.117187499999998, 3.858033323599208, 3.754537399884465, 3.4804390640125513, 3.4806162946866497]}
EXPECTED_WEIGHTED = [1.559731543624162, 1.8372153886417135, 2.177687005361243, 2.5236671649227707, 2.948187788859158]

_csv_cache = None


def workbook(data_only=True):
    return load_workbook(EXCEL_FILE, data_only=data_only)


def task_sheet(wb):
    return wb["Task"]


def find_task_csv():
    files = sorted(glob.glob(CSV_PATTERN))
    if not files:
        return None
    wb = workbook(data_only=False)
    idx = wb.sheetnames.index("Task")
    wb.close()
    candidate = f"/root/output/sheet.csv.{idx}"
    return candidate if Path(candidate).exists() else files[0]


def load_csv():
    global _csv_cache
    if _csv_cache is not None:
        return _csv_cache
    _csv_cache = {}
    csv_file = find_task_csv()
    if csv_file is None:
        return _csv_cache
    with open(csv_file, encoding="utf-8", errors="ignore") as handle:
        reader = csv.reader(handle)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                ref = f"{chr(ord('A') + col_idx - 1)}{row_idx}"
                if value in (None, ""):
                    _csv_cache[ref] = None
                    continue
                try:
                    _csv_cache[ref] = float(value)
                except ValueError:
                    _csv_cache[ref] = value
    return _csv_cache


def value(ws, ref):
    direct = ws[ref].value
    if isinstance(direct, (int, float)):
        return float(direct)
    cached = load_csv().get(ref)
    if isinstance(cached, (int, float)):
        return float(cached)
    return direct


def assert_matrix(ws, expected_map, label):
    errors = []
    for row, expected_values in expected_map.items():
        for idx, col in enumerate(COLUMNS):
            ref = f"{col}{row}"
            actual = value(ws, ref)
            expected = expected_values[idx]
            if not isinstance(actual, (int, float)) or abs(actual - expected) > TOLERANCE:
                errors.append(f"{ref}: expected {expected}, got {actual}")
    assert not errors, f"{label} mismatches:\n" + "\n".join(errors)


def test_file_and_sheet_structure():
    assert EXCEL_FILE.exists(), f"Missing output workbook: {EXCEL_FILE}"
    wb = workbook()
    assert wb.sheetnames == ["Task", "Data"]
    ws = task_sheet(wb)
    assert ws["A1"].value is not None
    wb.close()


def test_lookup_blocks():
    wb = workbook()
    ws = task_sheet(wb)
    assert_matrix(ws, EXPECTED_TOP, "top block")
    assert_matrix(ws, EXPECTED_MID, "middle block")
    assert_matrix(ws, EXPECTED_BASE, "base block")
    wb.close()


def test_derived_values_and_stats():
    wb = workbook()
    ws = task_sheet(wb)
    assert_matrix(ws, EXPECTED_NET, "net metric")
    assert_matrix(ws, EXPECTED_STATS, "statistics")
    weighted_errors = []
    for idx, col in enumerate(COLUMNS):
        ref = f"{col}50"
        actual = value(ws, ref)
        expected = EXPECTED_WEIGHTED[idx]
        if not isinstance(actual, (int, float)) or abs(actual - expected) > TOLERANCE:
            weighted_errors.append(f"{ref}: expected {expected}, got {actual}")
    wb.close()
    assert not weighted_errors, "weighted mean mismatches:\n" + "\n".join(weighted_errors)


def test_formulas_present_in_editable_ranges():
    wb = workbook(data_only=False)
    ws = task_sheet(wb)
    missing = []
    for row in list(range(12, 18)) + list(range(19, 25)) + list(range(26, 32)) + list(range(35, 41)) + [42, 43, 44, 45, 46, 47, 50]:
        for col in COLUMNS:
            ref = f"{col}{row}"
            cell_value = ws[ref].value
            if not (isinstance(cell_value, str) and cell_value.startswith("=")):
                missing.append(f"{ref}: {cell_value}")
    for col in COLUMNS:
        formula = ws[f"{col}50"].value
        if "SUMPRODUCT" not in str(formula).upper():
            missing.append(f"{col}50 missing SUMPRODUCT: {formula}")
    wb.close()
    assert not missing, "Missing formulas:\n" + "\n".join(missing)


def test_template_formatting_preserved():
    wb = workbook(data_only=False)
    ws = task_sheet(wb)
    for ref in ["H12", "L31", "H35", "L47", "H50"]:
        fill = ws[ref].fill
        rgb = getattr(fill.fgColor, "rgb", None)
        assert fill.patternType == "solid", f"{ref} lost yellow fill"
        assert rgb in YELLOW_RGBS, f"{ref} fill changed: {rgb}"
    wb.close()


def test_no_excel_errors_or_macros():
    errors = []
    csv_values = load_csv()
    for ref, value_ in csv_values.items():
        if isinstance(value_, str) and any(token in value_ for token in ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]):
            errors.append(f"{ref}: {value_}")
    with zipfile.ZipFile(EXCEL_FILE, "r") as archive:
        macros = [name for name in archive.namelist() if "vbaProject" in name or name.endswith(".bin")]
    if macros:
        errors.append(f"Macros found: {macros}")
    assert not errors, "Validation errors:\n" + "\n".join(errors)
