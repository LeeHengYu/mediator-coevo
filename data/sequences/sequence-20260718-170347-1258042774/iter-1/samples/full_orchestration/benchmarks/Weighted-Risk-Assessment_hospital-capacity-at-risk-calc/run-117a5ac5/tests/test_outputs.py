import csv
import glob
import json
import subprocess
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook

EXCEL_FILE = Path("/root/output/result.xlsx")
CSV_PATTERN = "/root/output/sheet.csv.*"
COLUMNS = ["H", "I", "J", "K", "L"]
TOLERANCE = 0.02
YELLOW_RGBS = {"00FFF2CC", "FFF2CC", "00FFF2CC"}

_recalculated_path = None


def _has_cached_values(path: Path) -> bool:
    """Check if the workbook has cached values for formula cells."""
    wb = load_workbook(path, data_only=True)
    ws = wb["Task"]
    for ref in ["H12", "H35", "H50"]:
        val = ws[ref].value
        if not isinstance(val, (int, float)):
            wb.close()
            return False
    wb.close()
    return True


def _recalculate_with_ssconvert(src: Path) -> Path:
    """Use ssconvert to recalculate formulas and write to a temp file."""
    tmp = Path(tempfile.mkdtemp()) / "recalculated.xlsx"
    subprocess.run(
        ["ssconvert", str(src), str(tmp), "--recalculate"],
        capture_output=True,
        check=True,
    )
    return tmp


def _get_reliable_workbook_path() -> Path:
    """Return path to workbook with cached values, recalculating if needed."""
    global _recalculated_path
    if _recalculated_path is not None and _recalculated_path.exists():
        return _recalculated_path
    if _has_cached_values(EXCEL_FILE):
        return EXCEL_FILE
    _recalculated_path = _recalculate_with_ssconvert(EXCEL_FILE)
    return _recalculated_path

EXPECTED_TOP = {"12": [252.0, 258.4, 264.6, 270.5, 276.0], "13": [231.2, 236.0, 240.5, 244.7, 248.6], "14": [205.5, 211.0, 216.7, 222.6, 228.8], "15": [278.0, 284.5, 291.2, 298.1, 305.2], "16": [219.4, 224.1, 228.7, 233.3, 238.0], "17": [240.3, 246.0, 252.0, 258.2, 264.6]}
EXPECTED_MID = {"19": [261.4, 266.6, 271.2, 275.2, 278.7], "20": [222.6, 228.6, 234.7, 241.0, 247.4], "21": [208.8, 214.2, 219.8, 225.8, 232.0], "22": [289.7, 295.0, 300.0, 304.6, 308.8], "23": [214.2, 219.0, 223.9, 228.9, 234.1], "24": [246.1, 251.8, 257.7, 263.8, 270.0]}
EXPECTED_BASE = {"26": [310.0, 318.0, 326.0, 334.0, 342.0], "27": [285.0, 292.0, 299.0, 306.0, 313.0], "28": [260.0, 267.0, 274.0, 281.0, 288.0], "29": [345.0, 353.0, 361.0, 369.0, 377.0], "30": [248.0, 255.0, 262.0, 269.0, 276.0], "31": [295.0, 303.0, 311.0, 319.0, 327.0]}
EXPECTED_NET = {"35": [-3.032258064516122, -2.5786163522012724, -2.024539877300603, -1.4071856287425115, -0.789473684210523], "36": [3.017543859649121, 2.534246575342468, 1.9397993311036827, 1.2091503267973818, 0.38338658146964494], "37": [-1.2692307692307736, -1.1985018726591719, -1.131386861313877, -1.1387900355871947, -1.1111111111111072], "38": [-3.391304347826084, -2.974504249291785, -2.437673130193909, -1.7615176151761516, -0.9549071618037195], "39": [2.0967741935483937, 1.9999999999999976, 1.8320610687022836, 1.6356877323420096, 1.4130434782608716], "40": [-1.9661016949152486, -1.9141914191419178, -1.8327974276527295, -1.7554858934169348, -1.651376146788984]}
EXPECTED_STATS = {"42": [-3.391304347826084, -2.974504249291785, -2.437673130193909, -1.7615176151761516, -1.651376146788984], "43": [3.017543859649121, 2.534246575342468, 1.9397993311036827, 1.6356877323420096, 1.4130434782608716], "44": [-1.617666232073011, -1.5563466459005448, -1.4820921444833033, -1.2729878321648531, -0.8721904230071212], "45": [-0.7574294705484522, -0.6885945529919469, -0.6090894827758587, -0.5363568522972336, -0.45173967403063614], "46": [-2.7657189721159035, -2.4125101189364337, -1.9766042648886346, -1.6684108272483291, -1.0720601237842602], "47": [1.255272952853602, 1.2003745318352053, 1.0911990861982435, 0.6221652362012375, 0.09017151504960297]}
EXPECTED_WEIGHTED = [-0.9409064830751549, -0.8501118568232691, -0.741953082378614, -0.6336528221512264, -0.5096203848153917]

_csv_cache = None


def workbook(data_only=True):
    path = _get_reliable_workbook_path() if data_only else EXCEL_FILE
    return load_workbook(path, data_only=data_only)


def task_sheet(wb):
    return wb["Task"]


def find_task_csv():
    files = sorted(glob.glob(CSV_PATTERN))
    if not files:
        return None
    wb = workbook(data_only=False)
    idx = wb.sheetnames.index("Task")
    wb.close()
    candidate = f"/root/output/sheet.csv.{idx}"
    return candidate if Path(candidate).exists() else files[0]


def load_csv():
    global _csv_cache
    if _csv_cache is not None:
        return _csv_cache
    _csv_cache = {}
    csv_file = find_task_csv()
    if csv_file is None:
        return _csv_cache
    with open(csv_file, encoding="utf-8", errors="ignore") as handle:
        reader = csv.reader(handle)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                ref = f"{chr(ord('A') + col_idx - 1)}{row_idx}"
                if value in (None, ""):
                    _csv_cache[ref] = None
                    continue
                try:
                    _csv_cache[ref] = float(value)
                except ValueError:
                    _csv_cache[ref] = value
    return _csv_cache


def value(ws, ref):
    direct = ws[ref].value
    if isinstance(direct, (int, float)):
        return float(direct)
    cached = load_csv().get(ref)
    if isinstance(cached, (int, float)):
        return float(cached)
    return direct


def assert_matrix(ws, expected_map, label):
    errors = []
    for row, expected_values in expected_map.items():
        for idx, col in enumerate(COLUMNS):
            ref = f"{col}{row}"
            actual = value(ws, ref)
            expected = expected_values[idx]
            if not isinstance(actual, (int, float)) or abs(actual - expected) > TOLERANCE:
                errors.append(f"{ref}: expected {expected}, got {actual}")
    assert not errors, f"{label} mismatches:\n" + "\n".join(errors)


def test_file_and_sheet_structure():
    assert EXCEL_FILE.exists(), f"Missing output workbook: {EXCEL_FILE}"
    wb = workbook()
    assert wb.sheetnames == ["Task", "Data"]
    ws = task_sheet(wb)
    assert ws["A1"].value is not None
    wb.close()


def test_lookup_blocks():
    wb = workbook()
    ws = task_sheet(wb)
    assert_matrix(ws, EXPECTED_TOP, "top block")
    assert_matrix(ws, EXPECTED_MID, "middle block")
    assert_matrix(ws, EXPECTED_BASE, "base block")
    wb.close()


def test_derived_values_and_stats():
    wb = workbook()
    ws = task_sheet(wb)
    assert_matrix(ws, EXPECTED_NET, "net metric")
    assert_matrix(ws, EXPECTED_STATS, "statistics")
    weighted_errors = []
    for idx, col in enumerate(COLUMNS):
        ref = f"{col}50"
        actual = value(ws, ref)
        expected = EXPECTED_WEIGHTED[idx]
        if not isinstance(actual, (int, float)) or abs(actual - expected) > TOLERANCE:
            weighted_errors.append(f"{ref}: expected {expected}, got {actual}")
    wb.close()
    assert not weighted_errors, "weighted mean mismatches:\n" + "\n".join(weighted_errors)


def test_formulas_present_in_editable_ranges():
    wb = workbook(data_only=False)
    ws = task_sheet(wb)
    missing = []
    for row in list(range(12, 18)) + list(range(19, 25)) + list(range(26, 32)) + list(range(35, 41)) + [42, 43, 44, 45, 46, 47, 50]:
        for col in COLUMNS:
            ref = f"{col}{row}"
            cell_value = ws[ref].value
            if not (isinstance(cell_value, str) and cell_value.startswith("=")):
                missing.append(f"{ref}: {cell_value}")
    for col in COLUMNS:
        formula = ws[f"{col}50"].value
        if "SUMPRODUCT" not in str(formula).upper():
            missing.append(f"{col}50 missing SUMPRODUCT: {formula}")
    wb.close()
    assert not missing, "Missing formulas:\n" + "\n".join(missing)


def test_template_formatting_preserved():
    wb = workbook(data_only=False)
    ws = task_sheet(wb)
    for ref in ["H12", "L31", "H35", "L47", "H50"]:
        fill = ws[ref].fill
        rgb = getattr(fill.fgColor, "rgb", None)
        assert fill.patternType == "solid", f"{ref} lost yellow fill"
        assert rgb in YELLOW_RGBS, f"{ref} fill changed: {rgb}"
    wb.close()


def test_no_excel_errors_or_macros():
    errors = []
    csv_values = load_csv()
    for ref, value_ in csv_values.items():
        if isinstance(value_, str) and any(token in value_ for token in ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]):
            errors.append(f"{ref}: {value_}")
    with zipfile.ZipFile(EXCEL_FILE, "r") as archive:
        macros = [name for name in archive.namelist() if "vbaProject" in name or name.endswith(".bin")]
    if macros:
        errors.append(f"Macros found: {macros}")
    assert not errors, "Validation errors:\n" + "\n".join(errors)
