import csv
import glob
import subprocess
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook

EXCEL_FILE = Path("/root/output/result.xlsx")
CSV_PATTERN = "/root/output/sheet.csv.*"
COLUMNS = ["H", "I", "J", "K", "L"]
TOLERANCE = 0.02
YELLOW_RGBS = {"00FFF2CC", "FFF2CC", "00FFF2CC"}

_recalculated_path = None


def _has_cached_values(path: Path) -> bool:
    """Check if the workbook has cached values for formula cells."""
    wb = load_workbook(path, data_only=True)
    ws = wb["Task"]
    for ref in ["H12", "H35", "H50"]:
        val = ws[ref].value
        if not isinstance(val, (int, float)):
            wb.close()
            return False
    wb.close()
    return True


def _recalculate_with_ssconvert(src: Path) -> Path:
    """Use ssconvert to recalculate formulas and write to a temp file."""
    tmp = Path(tempfile.mkdtemp()) / "recalculated.xlsx"
    try:
        subprocess.run(
            ["ssconvert", str(src), str(tmp), "--recalculate"],
            capture_output=True,
            check=True,
        )
    except subprocess.CalledProcessError as exc:
        stderr = (exc.stderr or b"").decode(errors="ignore")
        if "Unknown option --recalculate" not in stderr:
            raise
        subprocess.run(["ssconvert", str(src), str(tmp)], capture_output=True, check=True)
    return tmp


def _get_reliable_workbook_path() -> Path:
    """Return path to workbook with cached values, recalculating if needed."""
    global _recalculated_path
    if _recalculated_path is not None and _recalculated_path.exists():
        return _recalculated_path
    if _has_cached_values(EXCEL_FILE):
        return EXCEL_FILE
    _recalculated_path = _recalculate_with_ssconvert(EXCEL_FILE)
    return _recalculated_path

EXPECTED_TOP = {"12": [472.4, 488.6, 507.1, 528.3, 550.8], "13": [398.2, 409.8, 421.1, 434.5, 448.3], "14": [365.0, 379.4, 395.6, 409.1, 422.7], "15": [420.5, 438.7, 455.4, 473.8, 492.9], "16": [301.3, 314.0, 326.8, 340.4, 355.1], "17": [512.0, 528.9, 544.1, 559.4, 574.8]}
EXPECTED_MID = {"19": [520.0, 532.6, 543.9, 554.9, 565.2], "20": [420.6, 426.2, 429.5, 431.9, 432.6], "21": [347.5, 355.8, 364.0, 372.3, 380.5], "22": [370.1, 381.0, 391.0, 401.1, 411.0], "23": [309.4, 317.2, 324.1, 330.2, 335.6], "24": [482.0, 500.1, 519.3, 539.2, 559.4]}
EXPECTED_BASE = {"26": [680.0, 702.0, 725.0, 748.0, 772.0], "27": [550.0, 566.0, 582.0, 599.0, 616.0], "28": [430.0, 448.0, 466.0, 484.0, 502.0], "29": [505.0, 525.0, 545.0, 566.0, 588.0], "30": [390.0, 405.0, 420.0, 436.0, 452.0], "31": [610.0, 632.0, 655.0, 678.0, 701.0]}
EXPECTED_NET = {"35": [-7.0000000000000036, -6.267806267806268, -5.0758620689655105, -3.5561497326203235, -1.8652849740932758], "36": [-4.072727272727279, -2.8975265017667806, -1.443298969072161, 0.4340567612687851, 2.548701298701297], "37": [4.069767441860465, 5.267857142857135, 6.781115879828331, 7.603305785123969, 8.40637450199203], "38": [9.980198019801975, 10.990476190476189, 11.816513761467885, 12.844522968197877, 13.928571428571423], "39": [-2.076923076923068, -0.7901234567901207, 0.6428571428571401, 2.339449541284401, 4.314159292035398], "40": [4.918032786885246, 4.556962025316449, 3.7862595419847436, 2.9793510324483674, 2.1968616262482135]}
EXPECTED_STATS = {"42": [-7.0000000000000036, -6.267806267806268, -5.0758620689655105, -3.5561497326203235, -1.8652849740932758], "43": [9.980198019801975, 10.990476190476189, 11.816513761467885, 12.844522968197877, 13.928571428571423], "44": [0.9964221824686985, 1.883419284263164, 2.2145583424209416, 2.659400286866384, 3.4314302953683473], "45": [0.9697246498162224, 1.809973188714434, 2.7512642146834048, 3.7740893926171792, 4.921563862242514], "46": [-3.5737762237762265, -2.3706757405226155, -0.9217599410898356, 0.910404956272689, 2.2848215443614843], "47": [4.705966450629051, 5.090133363471963, 6.032401795367434, 6.447317096955068, 7.383320699502872]}
EXPECTED_WEIGHTED = [0.6255924170616096, 1.4185478950579604, 2.3076923076923115, 3.301053830817428, 4.414761773616079]

_csv_cache = None


def workbook(data_only=True):
    path = _get_reliable_workbook_path() if data_only else EXCEL_FILE
    return load_workbook(path, data_only=data_only)


def task_sheet(wb):
    return wb["Task"]


def find_task_csv():
    files = sorted(glob.glob(CSV_PATTERN))
    if not files:
        return None
    wb = workbook(data_only=False)
    idx = wb.sheetnames.index("Task")
    wb.close()
    candidate = f"/root/output/sheet.csv.{idx}"
    return candidate if Path(candidate).exists() else files[0]


def load_csv():
    global _csv_cache
    if _csv_cache is not None:
        return _csv_cache
    _csv_cache = {}
    csv_file = find_task_csv()
    if csv_file is None:
        return _csv_cache
    with open(csv_file, encoding="utf-8", errors="ignore") as handle:
        reader = csv.reader(handle)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                ref = f"{chr(ord('A') + col_idx - 1)}{row_idx}"
                if value in (None, ""):
                    _csv_cache[ref] = None
                    continue
                try:
                    _csv_cache[ref] = float(value)
                except ValueError:
                    _csv_cache[ref] = value
    return _csv_cache


def value(ws, ref):
    direct = ws[ref].value
    if isinstance(direct, (int, float)):
        return float(direct)
    cached = load_csv().get(ref)
    if isinstance(cached, (int, float)):
        return float(cached)
    return direct


def assert_matrix(ws, expected_map, label):
    errors = []
    for row, expected_values in expected_map.items():
        for idx, col in enumerate(COLUMNS):
            ref = f"{col}{row}"
            actual = value(ws, ref)
            expected = expected_values[idx]
            if not isinstance(actual, (int, float)) or abs(actual - expected) > TOLERANCE:
                errors.append(f"{ref}: expected {expected}, got {actual}")
    assert not errors, f"{label} mismatches:\n" + "\n".join(errors)


def test_file_and_sheet_structure():
    assert EXCEL_FILE.exists(), f"Missing output workbook: {EXCEL_FILE}"
    wb = workbook()
    assert wb.sheetnames == ["Task", "Data"]
    ws = task_sheet(wb)
    assert ws["A1"].value is not None
    wb.close()


def test_lookup_blocks():
    wb = workbook()
    ws = task_sheet(wb)
    assert_matrix(ws, EXPECTED_TOP, "top block")
    assert_matrix(ws, EXPECTED_MID, "middle block")
    assert_matrix(ws, EXPECTED_BASE, "base block")
    wb.close()


def test_derived_values_and_stats():
    wb = workbook()
    ws = task_sheet(wb)
    assert_matrix(ws, EXPECTED_NET, "net metric")
    assert_matrix(ws, EXPECTED_STATS, "statistics")
    weighted_errors = []
    for idx, col in enumerate(COLUMNS):
        ref = f"{col}50"
        actual = value(ws, ref)
        expected = EXPECTED_WEIGHTED[idx]
        if not isinstance(actual, (int, float)) or abs(actual - expected) > TOLERANCE:
            weighted_errors.append(f"{ref}: expected {expected}, got {actual}")
    wb.close()
    assert not weighted_errors, "weighted mean mismatches:\n" + "\n".join(weighted_errors)


def test_formulas_present_in_editable_ranges():
    wb = workbook(data_only=False)
    ws = task_sheet(wb)
    missing = []
    for row in list(range(12, 18)) + list(range(19, 25)) + list(range(26, 32)) + list(range(35, 41)) + [42, 43, 44, 45, 46, 47, 50]:
        for col in COLUMNS:
            ref = f"{col}{row}"
            cell_value = ws[ref].value
            if not (isinstance(cell_value, str) and cell_value.startswith("=")):
                missing.append(f"{ref}: {cell_value}")
    for col in COLUMNS:
        formula = ws[f"{col}50"].value
        if "SUMPRODUCT" not in str(formula).upper():
            missing.append(f"{col}50 missing SUMPRODUCT: {formula}")
    wb.close()
    assert not missing, "Missing formulas:\n" + "\n".join(missing)


def test_template_formatting_preserved():
    wb = workbook(data_only=False)
    ws = task_sheet(wb)
    for ref in ["H12", "L31", "H35", "L47", "H50"]:
        fill = ws[ref].fill
        rgb = getattr(fill.fgColor, "rgb", None)
        assert fill.patternType == "solid", f"{ref} lost yellow fill"
        assert rgb in YELLOW_RGBS, f"{ref} fill changed: {rgb}"
    wb.close()


def test_no_excel_errors_or_macros():
    errors = []
    csv_values = load_csv()
    for ref, value_ in csv_values.items():
        if isinstance(value_, str) and any(token in value_ for token in ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]):
            errors.append(f"{ref}: {value_}")
    with zipfile.ZipFile(EXCEL_FILE, "r") as archive:
        macros = [name for name in archive.namelist() if "vbaProject" in name or name.endswith(".bin")]
    if macros:
        errors.append(f"Macros found: {macros}")
    assert not errors, "Validation errors:\n" + "\n".join(errors)
