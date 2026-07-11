import re
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook

PLAN_FILE = Path("/root/assembly_plan.xlsx")
SUMMARY_FILE = Path("/root/assembly_summary.txt")
REFERENCE_FILE = Path("/root/assembly_schedule.xlsx")

PHASE_START = 6
PHASE_END = 54
INITIAL_PHASE6_TOTAL = 469.59
RATE_PER_DAY = 20.0


def _clean_label(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def load_reference_demands() -> Dict[int, float]:
    """
    Read the Assembly sheet with deduplication.
    The source file has duplicate phase entries and extra columns.
    We use FIRST occurrence and only the demand column (index 1).
    """
    wb = load_workbook(REFERENCE_FILE, data_only=True)
    ws = wb["Assembly"] if "Assembly" in wb.sheetnames else wb.active

    header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    phase_col_idx = None
    demand_col_idx = None
    for idx, val in enumerate(header):
        label = _clean_label(val)
        if label == "phase":
            phase_col_idx = idx
        elif label == "pcb assembly demand (std hrs)":
            demand_col_idx = idx

    assert phase_col_idx is not None, "Missing 'Phase' column in reference workbook"
    assert demand_col_idx is not None, "Missing 'PCB Assembly Demand (Std Hrs)' column in reference workbook"

    # FIRST-occurrence deduplication
    out: Dict[int, float] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        phase_val = row[phase_col_idx]
        demand_val = row[demand_col_idx]
        if phase_val is None or demand_val is None:
            continue
        try:
            phase = int(round(float(phase_val)))
            demand = float(demand_val)
        except (TypeError, ValueError):
            continue
        if phase not in out:
            out[phase] = demand

    for phase in range(PHASE_START, PHASE_END + 1):
        assert phase in out, f"Reference demand missing for phase {phase}"
    return out


def load_plan_rows() -> List[Tuple[int, int, float, float, float, float, float]]:
    assert PLAN_FILE.exists(), f"Output workbook missing: {PLAN_FILE}"
    wb = load_workbook(PLAN_FILE, data_only=True)
    ws = wb["Plan"] if "Plan" in wb.sheetnames else wb.active

    expected_headers = [
        "Phase",
        "Days Worked",
        "Scheduled Demand (Std Hrs)",
        "Weekly Capacity (Std Hrs)",
        "Start of Phase Past Due (Std Hrs)",
        "End of Phase Backlog/Buffer (Std Hrs)",
        "Overtime Hours",
    ]
    actual_headers = [ws.cell(1, i).value for i in range(1, 8)]
    assert actual_headers == expected_headers, (
        f"Header mismatch.\nExpected: {expected_headers}\nActual:   {actual_headers}"
    )

    rows: List[Tuple[int, int, float, float, float, float, float]] = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 8)]
        if all(v is None for v in vals):
            continue
        phase = int(round(float(vals[0])))
        days = int(round(float(vals[1])))
        demand = float(vals[2])
        capacity = float(vals[3])
        start_past_due = float(vals[4])
        end_backlog_buffer = float(vals[5])
        overtime = float(vals[6])
        rows.append((phase, days, demand, capacity, start_past_due, end_backlog_buffer, overtime))

    assert len(rows) == 49, f"Expected 49 phase rows (phases 6..54), found {len(rows)}"
    return rows


def expected_days(calc_start: float, demand: float) -> int:
    start_past_due = max(0.0, calc_start)
    if start_past_due > 0.01:
        if calc_start + demand - (RATE_PER_DAY * 5) <= 0.0:
            return 5
        return 6
    return 4 if demand <= 80.0 else 5


def test_plan_math_and_policy():
    ref_demands = load_reference_demands()
    rows = load_plan_rows()

    expected_phases = list(range(PHASE_START, PHASE_END + 1))
    actual_phases = [r[0] for r in rows]
    assert actual_phases == expected_phases, "Phases must be continuous from 6 to 54 with no gaps"

    calc_start = INITIAL_PHASE6_TOTAL - ref_demands[PHASE_START]
    prev_end = None
    for idx, (phase, days, demand, capacity, start_past_due, end_backlog_buffer, overtime) in enumerate(rows):
        assert days in {4, 5, 6}, f"Phase {phase}: days must be one of 4/5/6"

        # Demand must match source data (first-occurrence deduplicated).
        assert abs(demand - ref_demands[phase]) <= 0.01, (
            f"Phase {phase}: demand {demand} does not match reference {ref_demands[phase]}"
        )

        # Phase 6 initial condition.
        if phase == PHASE_START:
            assert abs(start_past_due + demand - INITIAL_PHASE6_TOTAL) <= 0.01, (
                "Phase 6 must satisfy Start of Phase Past Due + Scheduled Demand = 469.59"
            )

        # Carryover consistency.
        if prev_end is not None:
            assert abs(calc_start - prev_end) <= 0.01, f"Phase {phase}: calc carryover mismatch"

        # Start of phase past due should be max(0, signed start).
        assert abs(start_past_due - max(0.0, calc_start)) <= 0.01, (
            f"Phase {phase}: Start of Phase Past Due should equal max(0, prior end)"
        )

        # Capacity and backlog equations.
        assert abs(capacity - RATE_PER_DAY * days) <= 0.01, f"Phase {phase}: capacity formula mismatch"
        expected_end = calc_start + demand - capacity
        assert abs(end_backlog_buffer - expected_end) <= 0.01, f"Phase {phase}: end backlog formula mismatch"

        # Overtime formula.
        expected_ot = 10.0 * max(0, days - 4)
        assert abs(overtime - expected_ot) <= 0.01, f"Phase {phase}: overtime formula mismatch"

        # Deterministic day-selection policy.
        exp_days = expected_days(calc_start, demand)
        assert days == exp_days, f"Phase {phase}: expected {exp_days} day(s), got {days}"

        prev_end = end_backlog_buffer
        calc_start = end_backlog_buffer


def test_summary_file():
    assert SUMMARY_FILE.exists(), f"Summary file missing: {SUMMARY_FILE}"
    rows = load_plan_rows()

    first_phase_5 = next((phase for phase, days, *_ in rows if days == 5), None)
    first_phase_4 = next((phase for phase, days, *_ in rows if days == 4), None)

    content = SUMMARY_FILE.read_text(encoding="utf-8").strip()
    lines = content.splitlines()
    assert len(lines) == 3, "Summary file must contain exactly 3 lines"

    m5 = re.fullmatch(r"First_Week_5_Days:\s*(\d+|N/A)", lines[0].strip())
    m4 = re.fullmatch(r"First_Week_4_Days:\s*(\d+|N/A)", lines[1].strip())
    assert m5, "Line 1 must match: First_Week_5_Days: <phase-number-or-N/A>"
    assert m4, "Line 2 must match: First_Week_4_Days: <phase-number-or-N/A>"

    got_5 = m5.group(1)
    got_4 = m4.group(1)
    exp_5 = str(first_phase_5) if first_phase_5 is not None else "N/A"
    exp_4 = str(first_phase_4) if first_phase_4 is not None else "N/A"
    assert got_5 == exp_5, f"First_Week_5_Days mismatch: expected {exp_5}, got {got_5}"
    assert got_4 == exp_4, f"First_Week_4_Days mismatch: expected {exp_4}, got {got_4}"

    assert lines[2].startswith("Summary: "), "Line 3 must start with 'Summary: '"
    summary_text = lines[2][len("Summary: "):].strip()
    assert summary_text, "Summary text cannot be empty"

    words = re.findall(r"\S+", summary_text)
    assert len(words) <= 60, f"Summary exceeds 60 words ({len(words)})"
    sentence_count = len(re.findall(r"[.!?]", summary_text))
    assert sentence_count <= 3, f"Summary exceeds 3 sentences ({sentence_count})"

    # Must mention both step-down phase numbers or N/A strings.
    assert got_5 in summary_text, "Summary must mention the 5-day step-down phase"
    assert got_4 in summary_text, "Summary must mention the 4-day step-down phase"
