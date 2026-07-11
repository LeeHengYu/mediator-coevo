import re
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook

PLAN_FILE = Path("/root/dye_catch_up_plan.xlsx")
SUMMARY_FILE = Path("/root/dye_catch_up_summary.txt")
REFERENCE_FILE = Path("/root/dye_demand_sheet.xlsx")

WEEK_START = 3
WEEK_END = 51
INITIAL_WEEK3_TOTAL = 598.24
RATE_PER_DAY = 18.0


def _clean_label(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _read_sheet(ws, row_label: str) -> Dict[int, float]:
    """Read a sheet and return {week: value} from the row matching row_label."""
    week_col = None
    val_col = None
    header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    for idx, val in enumerate(header):
        if _clean_label(val) == "week":
            week_col = idx
        elif _clean_label(val) == row_label:
            val_col = idx
    if week_col is None:
        raise ValueError(f"Could not find 'Week' column in sheet")
    if val_col is None:
        raise ValueError(f"Could not find row labeled '{row_label}'")

    out: Dict[int, float] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        wv = row[week_col]
        vv = row[val_col]
        if wv is None or vv is None:
            continue
        try:
            week = int(round(float(wv)))
            val = float(vv)
        except (TypeError, ValueError):
            continue
        out[week] = val
    return out


def load_reference_demands() -> Dict[int, float]:
    """
    Read both sheets and compute effective demand = Dye Demand + Adjustment.
    The test independently computes what the effective demand should be.
    """
    wb = load_workbook(REFERENCE_FILE, data_only=True)

    dye_ws = wb["Dye"] if "Dye" in wb.sheetnames else wb.active
    adjust_ws = wb["Adjust"] if "Adjust" in wb.sheetnames else wb.active

    dye_demands = _read_sheet(dye_ws, "dye demand (std hrs)")
    adjust_demands = _read_sheet(adjust_ws, "demand adjustment (std hrs)")

    # Effective demand = Dye + Adjustment
    all_weeks = set(dye_demands.keys()) & set(adjust_demands.keys())

    for week in range(WEEK_START, WEEK_END + 1):
        assert week in all_weeks, f"Reference effective demand missing for week {week}"

    effective = {w: dye_demands[w] + adjust_demands[w] for w in range(WEEK_START, WEEK_END + 1)}
    return effective


def load_plan_rows() -> List[Tuple[int, int, float, float, float, float, float]]:
    assert PLAN_FILE.exists(), f"Output workbook missing: {PLAN_FILE}"
    wb = load_workbook(PLAN_FILE, data_only=True)
    ws = wb["Plan"] if "Plan" in wb.sheetnames else wb.active

    expected_headers = [
        "Week",
        "Days Worked",
        "Scheduled Demand (Std Hrs)",
        "Weekly Capacity (Std Hrs)",
        "Start of Week Past Due (Std Hrs)",
        "End of Week Backlog/Buffer (Std Hrs)",
        "Overtime Hours",
    ]
    actual_headers = [ws.cell(1, i).value for i in range(1, 8)]
    assert actual_headers == expected_headers, (
        f"Header mismatch.\nExpected: {expected_headers}\nActual:   {actual_headers}"
    )

    rows: List[Tuple[int, int, float, float, float, float, float]] = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 8)]
        if all(v is None for v in vals):
            continue
        week = int(round(float(vals[0])))
        days = int(round(float(vals[1])))
        demand = float(vals[2])
        capacity = float(vals[3])
        start_past_due = float(vals[4])
        end_backlog_buffer = float(vals[5])
        overtime = float(vals[6])
        rows.append((week, days, demand, capacity, start_past_due, end_backlog_buffer, overtime))

    assert len(rows) == 49, f"Expected 49 weekly rows (weeks 3..51), found {len(rows)}"
    return rows


def expected_days(calc_start: float, demand: float) -> int:
    start_past_due = max(0.0, calc_start)
    if start_past_due > 0.01:
        if calc_start + demand - (RATE_PER_DAY * 5) <= 0.0:
            return 5
        return 6
    return 4 if demand <= 72.0 else 5


def test_plan_math_and_policy():
    ref_demands = load_reference_demands()
    rows = load_plan_rows()

    expected_weeks = list(range(WEEK_START, WEEK_END + 1))
    actual_weeks = [r[0] for r in rows]
    assert actual_weeks == expected_weeks, "Weeks must be continuous from 3 to 51 with no gaps"

    calc_start = INITIAL_WEEK3_TOTAL - ref_demands[WEEK_START]
    prev_end = None
    for idx, (week, days, demand, capacity, start_past_due, end_backlog_buffer, overtime) in enumerate(rows):
        assert days in {4, 5, 6}, f"Week {week}: days must be one of 4/5/6"

        # Demand must match effective demand (Dye + Adjustment).
        assert abs(demand - ref_demands[week]) <= 0.01, (
            f"Week {week}: demand {demand} does not match effective demand {ref_demands[week]}"
        )

        # Week 3 initial condition.
        if week == WEEK_START:
            assert abs(start_past_due + demand - INITIAL_WEEK3_TOTAL) <= 0.01, (
                "Week 3 must satisfy Start of Week Past Due + Scheduled Demand = 598.24"
            )

        # Carryover consistency.
        if prev_end is not None:
            assert abs(calc_start - prev_end) <= 0.01, f"Week {week}: calc carryover mismatch"

        # Start of week past due should be max(0, signed start).
        assert abs(start_past_due - max(0.0, calc_start)) <= 0.01, (
            f"Week {week}: Start of Week Past Due should equal max(0, prior end)"
        )

        # Capacity and backlog equations.
        assert abs(capacity - RATE_PER_DAY * days) <= 0.01, f"Week {week}: capacity formula mismatch"
        expected_end = calc_start + demand - capacity
        assert abs(end_backlog_buffer - expected_end) <= 0.01, f"Week {week}: end backlog formula mismatch"

        # Overtime formula.
        expected_ot = 10.0 * max(0, days - 4)
        assert abs(overtime - expected_ot) <= 0.01, f"Week {week}: overtime formula mismatch"

        # Deterministic day-selection policy.
        exp_days = expected_days(calc_start, demand)
        assert days == exp_days, f"Week {week}: expected {exp_days} day(s), got {days}"

        prev_end = end_backlog_buffer
        calc_start = end_backlog_buffer


def test_summary_file():
    assert SUMMARY_FILE.exists(), f"Summary file missing: {SUMMARY_FILE}"
    rows = load_plan_rows()

    first_week_5 = next((week for week, days, *_ in rows if days == 5), None)
    first_week_4 = next((week for week, days, *_ in rows if days == 4), None)

    content = SUMMARY_FILE.read_text(encoding="utf-8").strip()
    lines = content.splitlines()
    assert len(lines) == 3, "Summary file must contain exactly 3 lines"

    m5 = re.fullmatch(r"First_Week_5_Days:\s*(\d+|N/A)", lines[0].strip())
    m4 = re.fullmatch(r"First_Week_4_Days:\s*(\d+|N/A)", lines[1].strip())
    assert m5, "Line 1 must match: First_Week_5_Days: <week-number-or-N/A>"
    assert m4, "Line 2 must match: First_Week_4_Days: <week-number-or-N/A>"

    got_5 = m5.group(1)
    got_4 = m4.group(1)
    exp_5 = str(first_week_5) if first_week_5 is not None else "N/A"
    exp_4 = str(first_week_4) if first_week_4 is not None else "N/A"
    assert got_5 == exp_5, f"First_Week_5_Days mismatch: expected {exp_5}, got {got_5}"
    assert got_4 == exp_4, f"First_Week_4_Days mismatch: expected {exp_4}, got {got_4}"

    assert lines[2].startswith("Summary: "), "Line 3 must start with 'Summary: '"
    summary_text = lines[2][len("Summary: "):].strip()
    assert summary_text, "Summary text cannot be empty"

    words = re.findall(r"\S+", summary_text)
    assert len(words) <= 60, f"Summary exceeds 60 words ({len(words)})"
    sentence_count = len(re.findall(r"[.!?]", summary_text))
    assert sentence_count <= 3, f"Summary exceeds 3 sentences ({sentence_count})"

    # Must mention both step-down week numbers or N/A strings.
    assert got_5 in summary_text, "Summary must mention the 5-day step-down week"
    assert got_4 in summary_text, "Summary must mention the 4-day step-down week"
