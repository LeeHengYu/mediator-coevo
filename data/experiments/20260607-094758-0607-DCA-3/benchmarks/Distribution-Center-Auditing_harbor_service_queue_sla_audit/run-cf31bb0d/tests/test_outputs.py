"""Verification tests for harbor_service_queue_sla_audit."""

from collections import defaultdict
from pathlib import Path
import os

import pytest
from docx import Document
from openpyxl import load_workbook

ROOT_DIR = os.environ.get('TASK_ROOT', '/root')
SOURCE_XLSX = f"{ROOT_DIR}/Ticket_Queue.xlsx"
OUTPUT_XLSX = f"{ROOT_DIR}/Service_Queue_SLA_Audit.xlsx"
OUTPUT_DOCX = f"{ROOT_DIR}/Service_Queue_SLA_Brief.docx"

BASE_HEADERS = [
    'Ticket ID',
    'Queue',
    'Priority Tier',
    'Open Age Hours',
    'Owner',
    'Escalation Code',
    'Region',
    'Analyst',
]

FORMATTED_HEADERS = BASE_HEADERS + ['SLA Breach', 'Missing Escalation', 'Total Errors', 'Error Summary']
SUMMARY_HEADERS = ['Queue', 'Region', 'SLA Breaches', 'Missing Escalations', 'Total Errors']


@pytest.fixture(scope='module')
def source_and_rules():
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    tickets_ws = wb['Tickets']
    rules_ws = wb['SLA_Rules']
    headers = [tickets_ws.cell(1, c).value for c in range(1, 9)]
    assert headers == BASE_HEADERS
    rows = []
    for r in range(2, tickets_ws.max_row + 1):
        vals = [tickets_ws.cell(r, c).value for c in range(1, 9)]
        if all(v is None for v in vals):
            continue
        rows.append(dict(zip(BASE_HEADERS, vals)))
    rules = {}
    for r in range(2, rules_ws.max_row + 1):
        tier = rules_ws.cell(r, 1).value
        if tier is None:
            continue
        rules[str(tier)] = {
            'max_hours': float(rules_ws.cell(r, 2).value),
            'escalation_required': str(rules_ws.cell(r, 3).value or '').strip().upper() == 'Y',
        }
    assert rows and rules
    return rows, rules


@pytest.fixture(scope='module')
def output_wb():
    assert Path(OUTPUT_XLSX).exists(), f'Missing output workbook: {OUTPUT_XLSX}'
    return load_workbook(OUTPUT_XLSX, data_only=True)


def calc_flags(row, rules):
    rule = rules[str(row['Priority Tier'])]
    sla_breach = 1 if float(row['Open Age Hours']) > rule['max_hours'] else 0
    escalation_code = str(row['Escalation Code'] or '').strip()
    missing_escalation = 1 if rule['escalation_required'] and not escalation_code else 0
    total = sla_breach + missing_escalation
    if total == 0:
        summary = 'None'
    elif sla_breach and missing_escalation:
        summary = 'SLA Breach, Missing Escalation'
    elif sla_breach:
        summary = 'SLA Breach'
    else:
        summary = 'Missing Escalation'
    return sla_breach, missing_escalation, total, summary


def expected_summary_rows(rows, rules):
    agg = defaultdict(lambda: [0, 0, 0])
    queue_totals = defaultdict(int)
    total_sla = total_missing = total_errors = 0
    for row in rows:
        sla, missing, total, _ = calc_flags(row, rules)
        key = (str(row['Queue']), str(row['Region']))
        agg[key][0] += sla
        agg[key][1] += missing
        agg[key][2] += total
        queue_totals[str(row['Queue'])] += total
        total_sla += sla
        total_missing += missing
        total_errors += total
    result_rows = []
    for (queue, region), values in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1])):
        if values[2] > 0:
            result_rows.append((queue, region, values[0], values[1], values[2]))
    return result_rows, (total_sla, total_missing, total_errors), queue_totals


def test_required_sheets_exist(output_wb):
    assert {'RawData', 'Formatted Data', 'Summary'}.issubset(set(output_wb.sheetnames))


def test_rawdata_copies_source_exactly(output_wb, source_and_rules):
    rows, _ = source_and_rules
    ws = output_wb['RawData']
    headers = [ws.cell(1, c).value for c in range(1, 9)]
    assert headers == BASE_HEADERS
    actual = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 9)]
        if all(v is None for v in vals):
            continue
        actual.append(vals)
    expected = [[row[h] for h in BASE_HEADERS] for row in rows]
    assert actual == expected


def test_formatted_data_logic(output_wb, source_and_rules):
    rows, rules = source_and_rules
    ws = output_wb['Formatted Data']
    headers = [ws.cell(1, c).value for c in range(1, 13)]
    assert headers == FORMATTED_HEADERS
    for i, row in enumerate(rows, start=2):
        out = [ws.cell(i, c).value for c in range(1, 13)]
        assert out[:8] == [row[h] for h in BASE_HEADERS]
        expected_sla, expected_missing, expected_total, expected_summary = calc_flags(row, rules)
        assert int(out[8]) == expected_sla
        assert int(out[9]) == expected_missing
        assert int(out[10]) == expected_total
        assert str(out[11]).strip() == expected_summary


def test_summary_sheet_matches_expected(output_wb, source_and_rules):
    rows, rules = source_and_rules
    ws = output_wb['Summary']
    headers = [ws.cell(1, c).value for c in range(1, 6)]
    assert headers == SUMMARY_HEADERS
    expected_rows, expected_totals, _ = expected_summary_rows(rows, rules)
    actual_rows = []
    grand_total = None
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 6)]
        if all(v is None for v in vals):
            continue
        if str(vals[0]).strip() == 'Grand Total':
            grand_total = (str(vals[0]), str(vals[1]), int(vals[2]), int(vals[3]), int(vals[4]))
        else:
            actual_rows.append((str(vals[0]), str(vals[1]), int(vals[2]), int(vals[3]), int(vals[4])))
    assert actual_rows == expected_rows
    assert grand_total == ('Grand Total', '-', expected_totals[0], expected_totals[1], expected_totals[2])


def test_word_summary_content(source_and_rules):
    rows, rules = source_and_rules
    assert Path(OUTPUT_DOCX).exists(), f'Missing output docx: {OUTPUT_DOCX}'
    _, expected_totals, queue_totals = expected_summary_rows(rows, rules)
    doc = Document(OUTPUT_DOCX)
    text = '\n'.join(p.text for p in doc.paragraphs if p.text.strip()).lower()
    assert 'sla breach' in text
    assert 'missing escalation' in text
    for value in expected_totals:
        assert str(value) in text
    top_queues = [q for q, total in sorted(queue_totals.items(), key=lambda x: (-x[1], x[0])) if total > 0][:4]
    mentioned = sum(1 for q in top_queues if q.lower() in text)
    assert mentioned >= 2, 'Word summary must mention at least two high-priority queues'
    assert any(marker in text for marker in ['recommend', 'review', 'should', 'action', 'prioritize'])
