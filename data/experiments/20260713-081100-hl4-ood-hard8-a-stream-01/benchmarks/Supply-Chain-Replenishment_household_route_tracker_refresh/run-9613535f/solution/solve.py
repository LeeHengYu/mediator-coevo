#!/usr/bin/env python3
import math
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook


def normalize_text(value: Any) -> str:
    return str(value or '').strip().upper()


def to_number(value: Any) -> float:
    if value in (None, ''):
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(',', '').strip()
    return float(text) if text else 0.0


def parse_date(value: Any) -> date | None:
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%m/%d/%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def to_iso(value: date | None) -> str:
    return value.isoformat() if value else ''


def diff_days(start: date, end: date) -> int:
    return (end - start).days


def add_days(start: date, days: int) -> date:
    return start + timedelta(days=days)


def round_to(value: float, decimals: int = 4) -> float:
    factor = 10 ** decimals
    return round(value * factor) / factor


def main():
    input_root = Path(sys.argv[1])
    output_root = Path(sys.argv[2])
    tpl_wb = load_workbook(input_root / 'Household_Route_Template.xlsx')
    stock_wb = load_workbook(input_root / 'Household_Current_Stock.xlsx', data_only=True)
    queue_wb = load_workbook(input_root / 'Household_Dispatch_Queue.xlsx', data_only=True)
    detail_ws = tpl_wb['Coverage_Detail']
    action_ws = tpl_wb['Dispatch_Plan']
    pack_ws = tpl_wb['Pack Matrix']
    alias_ws = tpl_wb['Route Alias Map']
    stock_ws = stock_wb['Route Snapshot']
    queue_ws = queue_wb['Queue Export']

    as_of = parse_date(stock_ws['B1'].value)
    horizon = parse_date(stock_ws['D1'].value)
    planning_days = diff_days(as_of, horizon)

    inventory = []
    current_route = ''
    r = 3
    while r <= stock_ws.max_row:
        label = str(stock_ws.cell(row=r, column=1).value or '').strip()
        if label.upper().startswith('ROUTE '):
            current_route = label.split(' ', 1)[1].strip().upper()
            r += 2
            continue
        sku = normalize_text(stock_ws.cell(row=r, column=1).value)
        if current_route and sku and sku != 'SKU':
            units = to_number(stock_ws.cell(row=r, column=2).value)
            rate = to_number(stock_ws.cell(row=r, column=3).value)
            inventory.append((current_route, sku, units, rate))
        r += 1

    alias_map: Dict[str, str] = {}
    for r in range(2, alias_ws.max_row + 1):
        alias = normalize_text(alias_ws.cell(row=r, column=1).value)
        route = normalize_text(alias_ws.cell(row=r, column=2).value)
        if alias and route:
            alias_map[alias] = route

    load_size_by_key: Dict[Tuple[str, str], float] = {}
    for r in range(2, pack_ws.max_row + 1):
        route = normalize_text(pack_ws.cell(row=r, column=1).value)
        sku = normalize_text(pack_ws.cell(row=r, column=2).value)
        size = to_number(pack_ws.cell(row=r, column=3).value)
        if route and sku and size > 0:
            load_size_by_key[(route, sku)] = size

    latest_by_queue: Dict[str, Tuple[int, str, str, date | None, float, str]] = {}
    for r in range(2, queue_ws.max_row + 1):
        row_type = normalize_text(queue_ws.cell(row=r, column=1).value)
        qid = normalize_text(queue_ws.cell(row=r, column=2).value)
        rev = int(to_number(queue_ws.cell(row=r, column=3).value))
        alias = normalize_text(queue_ws.cell(row=r, column=4).value)
        sku = normalize_text(queue_ws.cell(row=r, column=5).value)
        ship_date = parse_date(queue_ws.cell(row=r, column=6).value)
        cases = to_number(queue_ws.cell(row=r, column=7).value)
        state = normalize_text(queue_ws.cell(row=r, column=8).value)
        if row_type != 'DISPATCH' or not qid:
            continue
        prev = latest_by_queue.get(qid)
        if prev is None or rev > prev[0]:
            latest_by_queue[qid] = (rev, alias, sku, ship_date, cases, state)

    inbound_by_key: Dict[Tuple[str, str], List[Tuple[date, float]]] = {}
    valid_states = {'APPROVED', 'RELEASED'}
    for rev, alias, sku, ship_date, cases, state in latest_by_queue.values():
        route = alias_map.get(alias, '')
        if not route or not sku or ship_date is None or state not in valid_states:
            continue
        inbound_by_key.setdefault((route, sku), []).append((ship_date, cases))

    detail_rows = []
    action_rows = []
    for route, sku, units, rate in inventory:
        inbound = inbound_by_key.get((route, sku), [])
        inbound_in_horizon = sum(c for d, c in inbound if d <= horizon)
        earliest = min((d for d, c in inbound if d <= horizon), default=None)
        current_doh = units / rate if rate > 0 else None
        proj_oos = add_days(as_of, int(current_doh)) if current_doh is not None else None
        delivered_doh = (units + inbound_in_horizon) / rate if rate > 0 else None
        remaining_demand = rate * planning_days
        additional = max(0.0, remaining_demand - units - inbound_in_horizon) if rate > 0 else 0.0
        load_size = load_size_by_key[(route, sku)]
        loads = math.ceil(additional / load_size) if additional > 0 else 0
        req_date = proj_oos if loads > 0 else None
        earlier = loads > 0 and (earliest is None or (req_date and req_date < earliest))
        detail_rows.append([
            route, sku, units, rate,
            round_to(current_doh) if current_doh is not None else '',
            to_iso(proj_oos),
            inbound_in_horizon,
            round_to(delivered_doh) if delivered_doh is not None else '',
            round_to(remaining_demand),
            round_to(additional),
            loads,
            to_iso(req_date),
            earlier,
        ])
        if loads > 0:
            action_rows.append([route, sku, to_iso(req_date), loads, round_to(additional), earlier])

    for ws in [detail_ws, action_ws]:
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None

    detail_ws['A1'] = 'Field'
    detail_ws['B1'] = 'Value'
    detail_ws['A2'] = 'AsOfDate'
    detail_ws['B2'] = to_iso(as_of)
    detail_ws['A3'] = 'HorizonEnd'
    detail_ws['B3'] = to_iso(horizon)
    detail_ws['A4'] = 'PlanningDays'
    detail_ws['B4'] = planning_days
    headers = ['Route', 'SKU', 'On_Hand_Cases', 'Daily_Demand_Cases_Per_Day', 'Current_Days_On_Hand', 'Projected_OOS_Date', 'Inbound_Cases_By_Horizon', 'Delivered_Days_On_Hand', 'Remaining_Demand_Cases', 'Additional_Cases_Needed', 'Loads_Required', 'Required_Delivery_Date', 'Earlier_Delivery_Required']
    for i, h in enumerate(headers, 1):
        detail_ws.cell(row=6, column=i, value=h)
    for r, row in enumerate(detail_rows, 7):
        for c, val in enumerate(row, 1):
            detail_ws.cell(row=r, column=c, value=val)

    action_headers = ['Route', 'SKU', 'Required_Delivery_Date', 'Loads_Required', 'Additional_Cases_Needed', 'Earlier_Delivery_Required']
    for i, h in enumerate(action_headers, 1):
        action_ws.cell(row=1, column=i, value=h)
    for r, row in enumerate(action_rows, 2):
        for c, val in enumerate(row, 1):
            action_ws.cell(row=r, column=c, value=val)

    output_file = output_root / 'household_route_tracker_refresh.xlsx'
    tpl_wb.save(output_file)

    print('Wrote', output_file)


if __name__ == '__main__':
    main()
