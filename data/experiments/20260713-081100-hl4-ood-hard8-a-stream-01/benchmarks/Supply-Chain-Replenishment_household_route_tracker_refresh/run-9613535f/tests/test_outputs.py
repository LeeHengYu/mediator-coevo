#!/usr/bin/env python3
import math
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def normalize_text(value: Any) -> str:
    return str(value or '').strip().upper()


def to_number(value: Any) -> float:
    if value in (None, ''):
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(',', '').strip()
    return float(text) if text else 0.0


def parse_date(value: Any) -> date | None:
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%m/%d/%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def to_iso(value: date | None) -> str:
    return value.isoformat() if value else ''


def diff_days(start: date, end: date) -> int:
    return (end - start).days


def add_days(start: date, days: int) -> date:
    return start + timedelta(days=days)


def to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    return normalize_text(value) in ('TRUE', 'T', 'YES', 'Y', '1')


def round_to(value: float, decimals: int = 4) -> float:
    factor = 10 ** decimals
    return round(value * factor) / factor


def assert_approx(actual: Any, expected: float, tolerance: float, msg: str):
    a = to_number(actual)
    assert abs(a - expected) <= tolerance, f'{msg}: actual={a}, expected={expected}'


def compute_expected(input_root: Path):

    tpl_wb = load_workbook(input_root / 'Household_Route_Template.xlsx', data_only=True)
    stock_wb = load_workbook(input_root / 'Household_Current_Stock.xlsx', data_only=True)
    queue_wb = load_workbook(input_root / 'Household_Dispatch_Queue.xlsx', data_only=True)
    pack_ws = tpl_wb['Pack Matrix']
    alias_ws = tpl_wb['Route Alias Map']
    stock_ws = stock_wb['Route Snapshot']
    queue_ws = queue_wb['Queue Export']
    as_of = parse_date(stock_ws['B1'].value)
    horizon = parse_date(stock_ws['D1'].value)
    planning_days = diff_days(as_of, horizon)
    inventory = []
    current_route = ''
    r = 3
    while r <= stock_ws.max_row:
        label = str(stock_ws.cell(row=r, column=1).value or '').strip()
        if label.upper().startswith('ROUTE '):
            current_route = label.split(' ', 1)[1].strip().upper()
            r += 2
            continue
        sku = normalize_text(stock_ws.cell(row=r, column=1).value)
        if current_route and sku and sku != 'SKU':
            inventory.append((current_route, sku, to_number(stock_ws.cell(row=r, column=2).value), to_number(stock_ws.cell(row=r, column=3).value)))
        r += 1
    alias_map = {}
    for r in range(2, alias_ws.max_row + 1):
        alias = normalize_text(alias_ws.cell(row=r, column=1).value)
        route = normalize_text(alias_ws.cell(row=r, column=2).value)
        if alias and route:
            alias_map[alias] = route
    pack = {}
    for r in range(2, pack_ws.max_row + 1):
        route = normalize_text(pack_ws.cell(row=r, column=1).value)
        sku = normalize_text(pack_ws.cell(row=r, column=2).value)
        size = to_number(pack_ws.cell(row=r, column=3).value)
        if route and sku and size > 0:
            pack[(route, sku)] = size
    latest = {}
    for r in range(2, queue_ws.max_row + 1):
        row_type = normalize_text(queue_ws.cell(row=r, column=1).value)
        qid = normalize_text(queue_ws.cell(row=r, column=2).value)
        rev = int(to_number(queue_ws.cell(row=r, column=3).value))
        alias = normalize_text(queue_ws.cell(row=r, column=4).value)
        sku = normalize_text(queue_ws.cell(row=r, column=5).value)
        ship_date = parse_date(queue_ws.cell(row=r, column=6).value)
        cases = to_number(queue_ws.cell(row=r, column=7).value)
        state = normalize_text(queue_ws.cell(row=r, column=8).value)
        if row_type != 'DISPATCH' or not qid:
            continue
        prev = latest.get(qid)
        if prev is None or rev > prev[0]:
            latest[qid] = (rev, alias, sku, ship_date, cases, state)
    inbound_by_key = {}
    valid = {'APPROVED', 'RELEASED'}
    for rev, alias, sku, ship_date, cases, state in latest.values():
        route = alias_map.get(alias, '')
        if not route or not sku or ship_date is None or state not in valid:
            continue
        inbound_by_key.setdefault((route, sku), []).append((ship_date, cases))
    detail_map = {}
    detail_order = []
    action_map = {}
    action_keys = set()
    for route, sku, units, rate in inventory:
        key = (route, sku)
        detail_order.append(key)
        inbound = inbound_by_key.get(key, [])
        inbound_in_horizon = sum(c for d, c in inbound if d <= horizon)
        earliest = min((d for d, c in inbound if d <= horizon), default=None)
        current_doh = units / rate if rate > 0 else None
        proj_oos = add_days(as_of, int(current_doh)) if current_doh is not None else None
        delivered_doh = (units + inbound_in_horizon) / rate if rate > 0 else None
        remaining = rate * planning_days
        additional = max(0.0, remaining - units - inbound_in_horizon) if rate > 0 else 0.0
        loads = math.ceil(additional / pack[(route, sku)]) if additional > 0 else 0
        req = proj_oos if loads > 0 else None
        earlier = loads > 0 and (earliest is None or (req and req < earliest))
        detail_map[key] = {'Route': route, 'SKU': sku, 'On_Hand_Cases': units, 'Daily_Demand_Cases_Per_Day': rate, 'Current_Days_On_Hand': round_to(current_doh) if current_doh is not None else '', 'Projected_OOS_Date': to_iso(proj_oos), 'Inbound_Cases_By_Horizon': inbound_in_horizon, 'Delivered_Days_On_Hand': round_to(delivered_doh) if delivered_doh is not None else '', 'Remaining_Demand_Cases': round_to(remaining), 'Additional_Cases_Needed': round_to(additional), 'Loads_Required': loads, 'Required_Delivery_Date': to_iso(req), 'Earlier_Delivery_Required': earlier}
        if loads > 0:
            action_keys.add(key)
            action_map[key] = {'Route': route, 'SKU': sku, 'Required_Delivery_Date': to_iso(req), 'Loads_Required': loads, 'Additional_Cases_Needed': round_to(additional), 'Earlier_Delivery_Required': earlier}
    return {'as_of': as_of, 'horizon': horizon, 'planning_days': planning_days, 'detail_order': detail_order, 'detail_map': detail_map, 'action_map': action_map, 'action_keys': action_keys}



def read_actual(output_file: Path):
    wb = load_workbook(output_file, data_only=True)
    assert set(wb.sheetnames) == set(['Overview', 'Coverage_Detail', 'Dispatch_Plan', 'Pack Matrix', 'Route Alias Map']), f'Sheet names mismatch: {wb.sheetnames}'
    detail_ws = wb['Coverage_Detail']
    action_ws = wb['Dispatch_Plan']

    assert wb['Overview']['A1'].value.strip().lower() == 'household route dispatch tracker'
    assert wb['Pack Matrix']['A1'].value.strip().lower() == 'route'
    assert wb['Route Alias Map']['A1'].value.strip().lower() == 'alias'

    assert detail_ws['A1'].value.strip().lower() == 'field'
    assert detail_ws['B1'].value.strip().lower() == 'value'
    as_of = parse_date(detail_ws['B2'].value)
    horizon = parse_date(detail_ws['B3'].value)
    planning_days = int(to_number(detail_ws['B4'].value))
    header = [detail_ws.cell(row=6, column=c).value for c in range(1, 13 + 1)]
    assert set(header) == set(['Route', 'SKU', 'On_Hand_Cases', 'Daily_Demand_Cases_Per_Day', 'Current_Days_On_Hand', 'Projected_OOS_Date', 'Inbound_Cases_By_Horizon', 'Delivered_Days_On_Hand', 'Remaining_Demand_Cases', 'Additional_Cases_Needed', 'Loads_Required', 'Required_Delivery_Date', 'Earlier_Delivery_Required']), f'Detail header mismatch: {header}'
    detail_order = []
    detail_map = {}
    for r in range(7, detail_ws.max_row + 1):
        row = [detail_ws.cell(row=r, column=c).value for c in range(1, 13 + 1)]
        key = tuple(normalize_text(v) for v in row[:2])
        if not all(key):
            continue
        detail_order.append(key)
        detail_map[key] = dict(zip(['Route', 'SKU', 'On_Hand_Cases', 'Daily_Demand_Cases_Per_Day', 'Current_Days_On_Hand', 'Projected_OOS_Date', 'Inbound_Cases_By_Horizon', 'Delivered_Days_On_Hand', 'Remaining_Demand_Cases', 'Additional_Cases_Needed', 'Loads_Required', 'Required_Delivery_Date', 'Earlier_Delivery_Required'], row))
    action_header = [action_ws.cell(row=1, column=c).value for c in range(1, 6 + 1)]
    assert set(action_header) == set(['Route', 'SKU', 'Required_Delivery_Date', 'Loads_Required', 'Additional_Cases_Needed', 'Earlier_Delivery_Required']), f'Action header mismatch: {action_header}'
    action_order = []
    action_map = {}
    for r in range(2, action_ws.max_row + 1):
        row = [action_ws.cell(row=r, column=c).value for c in range(1, 6 + 1)]
        key = tuple(normalize_text(v) for v in row[:2])
        if not all(key):
            continue
        action_order.append(key)
        action_map[key] = dict(zip(['Route', 'SKU', 'Required_Delivery_Date', 'Loads_Required', 'Additional_Cases_Needed', 'Earlier_Delivery_Required'], row))
    return {'as_of': as_of, 'horizon': horizon, 'planning_days': planning_days, 'detail_order': detail_order, 'detail_map': detail_map, 'action_order': action_order, 'action_map': action_map}


def compare(expected: dict, actual: dict):
    assert actual['as_of'] == expected['as_of']
    assert actual['horizon'] == expected['horizon']
    assert actual['planning_days'] == expected['planning_days']
    assert set(actual['detail_order']) == set(expected['detail_order'])
    for key in expected['detail_order']:
        erow = expected['detail_map'][key]
        arow = actual['detail_map'][key]
        for field in ['Route', 'SKU', 'On_Hand_Cases', 'Daily_Demand_Cases_Per_Day', 'Current_Days_On_Hand', 'Projected_OOS_Date', 'Inbound_Cases_By_Horizon', 'Delivered_Days_On_Hand', 'Remaining_Demand_Cases', 'Additional_Cases_Needed', 'Loads_Required', 'Required_Delivery_Date', 'Earlier_Delivery_Required']:
            ev = erow[field]
            av = arow[field]
            if field in ['On_Hand_Cases', 'Daily_Demand_Cases_Per_Day', 'Current_Days_On_Hand', 'Inbound_Cases_By_Horizon', 'Delivered_Days_On_Hand', 'Remaining_Demand_Cases', 'Additional_Cases_Needed', 'Loads_Required']:
                assert_approx(av, ev, 1e-4, f'detail {key} {field}')
            elif field in ['Projected_OOS_Date', 'Required_Delivery_Date']:
                assert normalize_text(av) == normalize_text(ev), f'detail {key} {field}: {av} vs {ev}'
            elif field in ['Earlier_Delivery_Required']:
                assert to_bool(av) == bool(ev), f'detail {key} {field}: {av} vs {ev}'
            else:
                assert normalize_text(av) == normalize_text(ev), f'detail {key} {field}: {av} vs {ev}'
    expected_action_order = [k for k in expected['detail_order'] if k in expected['action_keys']]
    assert set(actual['action_order']) == set(expected_action_order)
    for key in expected_action_order:
        erow = expected['action_map'][key]
        arow = actual['action_map'][key]
        for field in ['Route', 'SKU', 'Required_Delivery_Date', 'Loads_Required', 'Additional_Cases_Needed', 'Earlier_Delivery_Required']:
            ev = erow[field]
            av = arow[field]
            if field in ['On_Hand_Cases', 'Daily_Demand_Cases_Per_Day', 'Current_Days_On_Hand', 'Inbound_Cases_By_Horizon', 'Delivered_Days_On_Hand', 'Remaining_Demand_Cases', 'Additional_Cases_Needed', 'Loads_Required']:
                assert_approx(av, ev, 1e-4, f'action {key} {field}')
            elif field in ['Projected_OOS_Date', 'Required_Delivery_Date']:
                assert normalize_text(av) == normalize_text(ev), f'action {key} {field}: {av} vs {ev}'
            elif field in ['Earlier_Delivery_Required']:
                assert to_bool(av) == bool(ev), f'action {key} {field}: {av} vs {ev}'
            else:
                assert normalize_text(av) == normalize_text(ev), f'action {key} {field}: {av} vs {ev}'


def main():
    in_harness = Path('/tests').exists() and Path('/root').exists()
    task_dir = Path(__file__).resolve().parents[1]
    input_root = Path('/root') if in_harness else (task_dir / 'environment')
    output_file = Path('/root/household_route_tracker_refresh.xlsx') if in_harness else (task_dir / 'household_route_tracker_refresh.xlsx')
    assert output_file.exists(), f'Missing output file: {output_file}'
    expected = compute_expected(input_root)
    actual = read_actual(output_file)
    compare(expected, actual)
    print('All checks passed.')


if __name__ == '__main__':
    main()
