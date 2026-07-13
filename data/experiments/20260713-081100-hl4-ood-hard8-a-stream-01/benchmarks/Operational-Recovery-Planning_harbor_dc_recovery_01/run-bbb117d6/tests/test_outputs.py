from datetime import date, datetime, timedelta
from pathlib import Path
import re

import pytest
from openpyxl import load_workbook


OUTPUT_XLSX = "/root/server_provisioning_recovery_plan_analysis.xlsx"
OUTPUT_SUMMARY = "/root/server_provisioning_recovery_summary.md"
REFERENCE_XLSX = "/root/Open_Server_Requests_Listing.xlsx"

SHEET_NAMES = [
    "Current Capacity and Racks",
    "Relocated Network Equipment",
    "10 hr Shift Relocate Network Eq",
]

START_DATE = date(2018, 1, 22)
END_DATE = date(2018, 5, 1)
FIRST_ROW = 4
LAST_ROW = 103
HOLIDAYS = {date(2018, 2, 19), date(2018, 3, 30)}


def normalize_formula(value):
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text.startswith("="):
        return None
    return text.upper().replace("$", "").replace(" ", "")


def assert_formula_contains(value, tokens, cell_ref):
    formula = normalize_formula(value)
    assert formula is not None, f"{cell_ref} must be a formula."
    for token in tokens:
        assert token in formula, f"{cell_ref} formula is missing token '{token}'. Got: {value}"


def parse_nonnegative_int(value, cell_ref):
    if value is None or value == "":
        return 0
    if isinstance(value, bool):
        pytest.fail(f"{cell_ref} must be numeric, got boolean.")
    if isinstance(value, (int, float)):
        assert float(value).is_integer(), f"{cell_ref} must be a whole number."
        assert value >= 0, f"{cell_ref} must be nonnegative."
        return int(value)
    if isinstance(value, str):
        text = value.strip()
        assert not text.startswith("="), f"{cell_ref} must be a numeric constant, not a formula."
        if text == "":
            return 0
        try:
            num = float(text)
        except ValueError as exc:
            raise AssertionError(f"{cell_ref} must be numeric. Got: {value!r}") from exc
        assert num.is_integer(), f"{cell_ref} must be a whole number."
        assert num >= 0, f"{cell_ref} must be nonnegative."
        return int(num)
    pytest.fail(f"{cell_ref} has unsupported type: {type(value)!r}")


def as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def build_due_map_from_reference(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    a_open = {}
    b_open = {}
    for r in range(5, ws.max_row + 1):
        a_label = ws.cell(r, 2).value
        a_value = ws.cell(r, 5).value
        b_label = ws.cell(r, 7).value
        b_value = ws.cell(r, 10).value

        if isinstance(a_label, str):
            a_open[a_label.strip().lower()] = int(a_value or 0)
        if isinstance(b_label, str):
            b_open[b_label.strip().lower()] = int(b_value or 0)

    def a(label):
        return a_open[label.lower()]

    def b(label):
        return b_open[label.lower()]

    return {
        date(2018, 1, 22): (
            a("December SR143776") + a("January SR305419"),
            b("November SR029472") + b("December SR143776") + b("January SR305419"),
        ),
        date(2018, 2, 1): (a("February SR421530"), b("February SR421530")),
        date(2018, 2, 15): (a("February Sup SR470517"), b("February Sup SR470517")),
        date(2018, 3, 1): (a("Expected March"), b("Expected March")),
        date(2018, 4, 2): (a("Expected April"), b("Expected April")),
        date(2018, 5, 1): (a("Expected May"), b("Expected May")),
    }


@pytest.fixture(scope="module")
def workbook():
    output = Path(OUTPUT_XLSX)
    assert output.exists(), f"Output workbook not found: {OUTPUT_XLSX}"
    return load_workbook(output, data_only=False)


@pytest.fixture(scope="module")
def due_map():
    reference = Path(REFERENCE_XLSX)
    assert reference.exists(), f"Reference workbook missing: {REFERENCE_XLSX}"
    return build_due_map_from_reference(reference)


@pytest.fixture(scope="module")
def summary_text():
    summary = Path(OUTPUT_SUMMARY)
    assert summary.exists(), f"Summary markdown not found: {OUTPUT_SUMMARY}"
    return summary.read_text(encoding="utf-8")


def test_sheet_names(workbook):
    assert set(workbook.sheetnames) == set(SHEET_NAMES)


@pytest.mark.parametrize("sheet_name", SHEET_NAMES)
def test_sheet_layout_and_rules(workbook, due_map, sheet_name):
    ws = workbook[sheet_name]

    assert ws.max_row == LAST_ROW, f"{sheet_name}: expected exactly {LAST_ROW} rows."

    assert ws["C2"].value == "Rack-Mount Web Servers"
    assert ws["F2"].value == "Blade Database Servers"
    assert ws["I2"].value == "Network Appliances"

    expected_headers = [
        "Planned Production",
        "Purchase Orders Due",
        "Cumulative Open Purchase Orders (EOD)",
        "Planned Production",
        "Purchase Orders Due",
        "Cumulative Open Purchase Orders (EOD)",
        "Actual Var to PO",
        "Total Prod",
        "Notes",
    ]
    actual_headers = [ws.cell(3, col).value for col in range(3, 12)]
    assert set(actual_headers) == set(expected_headers), f"{sheet_name}: header mismatch in C3:K3."

    first_date = as_date(ws.cell(FIRST_ROW, 2).value)
    assert first_date == START_DATE, f"{sheet_name}: B4 must be {START_DATE}."

    total_c = 0
    c_days = 0
    c_before_feb1 = 0
    c_on_or_after_feb1 = 0
    first_b_date = None
    high_output_days = 0
    a_backlog = 0
    b_backlog = 0

    for row in range(FIRST_ROW, LAST_ROW + 1):
        d = START_DATE + timedelta(days=row - FIRST_ROW)

        if row > FIRST_ROW:
            b_value = ws.cell(row, 2).value
            literal = as_date(b_value)
            if literal is not None:
                assert literal == d, f"{sheet_name}: B{row} has incorrect date."
            else:
                formula = normalize_formula(b_value)
                assert formula is not None, f"{sheet_name}: B{row} must be date or formula."
                assert f"B{row - 1}" in formula and "+1" in formula, (
                    f"{sheet_name}: B{row} must advance date by 1 day."
                )

        a_plan = parse_nonnegative_int(ws.cell(row, 3).value, f"{sheet_name}!C{row}")
        a_due = parse_nonnegative_int(ws.cell(row, 4).value, f"{sheet_name}!D{row}")
        b_plan = parse_nonnegative_int(ws.cell(row, 6).value, f"{sheet_name}!F{row}")
        b_due = parse_nonnegative_int(ws.cell(row, 7).value, f"{sheet_name}!G{row}")
        c_plan = parse_nonnegative_int(ws.cell(row, 9).value, f"{sheet_name}!I{row}")

        assert_formula_contains(ws.cell(row, 5).value, [f"D{row}", f"C{row}"], f"{sheet_name}!E{row}")
        assert_formula_contains(ws.cell(row, 8).value, [f"G{row}", f"F{row}"], f"{sheet_name}!H{row}")
        assert_formula_contains(
            ws.cell(row, 10).value,
            [f"C{row}", f"F{row}", f"I{row}"],
            f"{sheet_name}!J{row}",
        )
        if row > FIRST_ROW:
            assert_formula_contains(ws.cell(row, 5).value, [f"E{row - 1}"], f"{sheet_name}!E{row}")
            assert_formula_contains(ws.cell(row, 8).value, [f"H{row - 1}"], f"{sheet_name}!H{row}")

        expected_a_due, expected_b_due = due_map.get(d, (0, 0))
        assert a_due == expected_a_due, (
            f"{sheet_name}: wrong Web due on {d}. Expected {expected_a_due}, got {a_due}."
        )
        assert b_due == expected_b_due, (
            f"{sheet_name}: wrong Database due on {d}. Expected {expected_b_due}, got {b_due}."
        )

        if d.weekday() >= 5 or d in HOLIDAYS:
            assert a_plan == 0, f"{sheet_name}: Web production must be 0 on non-working day {d}."
            assert b_plan == 0, f"{sheet_name}: Database production must be 0 on non-working day {d}."
            assert c_plan == 0, f"{sheet_name}: Network Appliance production must be 0 on non-working day {d}."

        if sheet_name in ("Current Capacity and Racks", "Relocated Network Equipment"):
            if d < date(2018, 2, 5):
                assert a_plan <= 120, f"{sheet_name}: Web output exceeds 120 on {d}."
                assert b_plan <= 120, f"{sheet_name}: Database output exceeds 120 on {d}."
            else:
                assert a_plan <= 135, f"{sheet_name}: Web output exceeds 135 on {d}."
                assert b_plan <= 135, f"{sheet_name}: Database output exceeds 135 on {d}."
        else:
            if a_plan > 135 or b_plan > 135:
                high_output_days += 1
                assert d >= date(2018, 2, 1), f"{sheet_name}: 10-hour output used before 2018-02-01."
                assert d.weekday() < 5 and d not in HOLIDAYS, (
                    f"{sheet_name}: high-output day must be a working day ({d})."
                )
                assert a_plan <= 170, f"{sheet_name}: Web output exceeds 170 on {d}."
                assert b_plan <= 170, f"{sheet_name}: Database output exceeds 170 on {d}."
            else:
                if d < date(2018, 2, 5):
                    assert a_plan <= 120, f"{sheet_name}: Web output exceeds 120 on {d}."
                    assert b_plan <= 120, f"{sheet_name}: Database output exceeds 120 on {d}."
                else:
                    assert a_plan <= 135, f"{sheet_name}: Web output exceeds 135 on {d}."
                    assert b_plan <= 135, f"{sheet_name}: Database output exceeds 135 on {d}."

        if b_plan > 0 and first_b_date is None:
            first_b_date = d

        total_c += c_plan
        if c_plan > 0:
            c_days += 1
        if d < date(2018, 2, 1):
            c_before_feb1 += c_plan
        else:
            c_on_or_after_feb1 += c_plan

        a_backlog += a_due - a_plan
        b_backlog += b_due - b_plan

    if sheet_name == "Current Capacity and Racks":
        assert total_c >= 1200, f"{sheet_name}: total Network Appliances output must be >= 1200."
        assert c_days >= 10, f"{sheet_name}: expected Network Appliances to run on at least 10 days."
        assert first_b_date is not None and first_b_date >= date(2018, 3, 1), (
            f"{sheet_name}: Blade Database Servers production starts too early."
        )
        assert a_backlog > 0 and b_backlog > 0, (
            f"{sheet_name}: expected both Web and Database backlogs to remain > 0 by May 1."
        )

    if sheet_name == "Relocated Network Equipment":
        assert c_before_feb1 >= 100, f"{sheet_name}: must produce at least 100 Network Appliances units pre-relocation."
        assert c_on_or_after_feb1 == 0, f"{sheet_name}: Network Appliances output must be 0 on/after 2018-02-01."
        assert first_b_date is not None and first_b_date >= date(2018, 2, 20), (
            f"{sheet_name}: Blade Database Servers production starts too early."
        )
        assert a_backlog <= 0 and b_backlog > 0, (
            f"{sheet_name}: expected Web backlog cleared and Database backlog remaining by May 1."
        )

    if sheet_name == "10 hr Shift Relocate Network Eq":
        assert total_c == 0, f"{sheet_name}: Network Appliances output must be zero for entire horizon."
        assert 20 <= high_output_days <= 24, (
            f"{sheet_name}: expected 20-24 high-output (10-hour shift) days, got {high_output_days}."
        )
        assert first_b_date is not None and first_b_date >= date(2018, 2, 20), (
            f"{sheet_name}: Blade Database Servers production starts too early."
        )
        assert a_backlog <= 0 and b_backlog <= 0, (
            f"{sheet_name}: expected both Web and Database backlogs cleared by May 1."
        )


def extract_section(text, scenario_number):
    pattern = rf"##\s*Scenario\s*{scenario_number}\b(.*?)(?=\n##\s*Scenario\s*[123]\b|\Z)"
    match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
    assert match, f"Missing section header: ## Scenario {scenario_number}"
    return match.group(1)


def test_summary_structure_and_required_statements(summary_text):
    sec1 = extract_section(summary_text, 1)
    sec2 = extract_section(summary_text, 2)
    sec3 = extract_section(summary_text, 3)

    required_labels = [
        "Actions:",
        "Rack-Mount Web Servers Impact:",
        "Blade Database Servers Impact:",
        "Network Appliances Impact:",
        "May PO On-Time:",
    ]

    for idx, section in enumerate([sec1, sec2, sec3], start=1):
        for label in required_labels:
            assert re.search(rf"^\s*{re.escape(label)}", section, flags=re.IGNORECASE | re.MULTILINE), (
                f"Scenario {idx} is missing required field '{label}'."
            )

    assert re.search(r"May PO On-Time:\s*No\b", sec1, flags=re.IGNORECASE), (
        "Scenario 1 must contain 'May PO On-Time: No'."
    )
    assert re.search(
        r"May PO On-Time:\s*Web\s+Yes,\s*Database\s+No\b",
        sec2,
        flags=re.IGNORECASE,
    ), "Scenario 2 must contain 'May PO On-Time: Web Yes, Database No'."
    assert re.search(r"May PO On-Time:\s*Yes\b", sec3, flags=re.IGNORECASE), (
        "Scenario 3 must contain 'May PO On-Time: Yes'."
    )
    assert re.search(r"30[- ]day notification", sec3, flags=re.IGNORECASE), (
        "Scenario 3 must explicitly mention '30-day notification'."
    )
