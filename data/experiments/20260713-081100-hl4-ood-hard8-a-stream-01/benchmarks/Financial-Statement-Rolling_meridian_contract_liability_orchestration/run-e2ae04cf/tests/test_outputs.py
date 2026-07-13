from __future__ import annotations

import math
from pathlib import Path

from openpyxl import load_workbook

SPEC = {
  "sheet_order": [
    "Contract Liability Summary",
    "Subscriptions #2350",
    "Services #2355"
  ],
  "summary": {
    "sheet_name": "Contract Liability Summary",
    "company": "Meridian Mobility Network",
    "title": "Contract Liability Summary",
    "period_text": "For the period ending 12/31/2025",
    "total_label": "Total contract liability balance at 12/31/2025"
  },
  "details": [
    {
      "sheet_name": "Subscriptions #2350",
      "company": "Meridian Mobility Network",
      "title": "Subscription Contract Liability #2350 as of 12/31/2025",
      "subtitle": "Contract liability activity from September through December 2025",
      "entity_header": "Customer",
      "term_header": "Contract Months",
      "notes_header": "Notes",
      "account_header": "Liability Account",
      "totals_label": "Period Totals",
      "gl_key": "subscriptions_2350",
      "months": [
        {
          "slug": "sep",
          "label": "Sep",
          "adds_subheader": "Billings",
          "release_subheader": "Revenue"
        },
        {
          "slug": "oct",
          "label": "Oct",
          "adds_subheader": "Billings",
          "release_subheader": "Revenue"
        },
        {
          "slug": "nov",
          "label": "Nov",
          "adds_subheader": "Billings",
          "release_subheader": "Revenue"
        },
        {
          "slug": "dec",
          "label": "Dec",
          "adds_subheader": "Billings",
          "release_subheader": "Revenue"
        }
      ],
      "summary_labels": {
        "section": "Subscription Contract Liability (Acct 2350)",
        "total": "Subscription billings in period",
        "ending": "Subscription revenue recognized in period",
        "gl": "Subscription GL balance at 12/31/2025"
      }
    },
    {
      "sheet_name": "Services #2355",
      "company": "Meridian Mobility Network",
      "title": "Services Contract Liability #2355 as of 12/31/2025",
      "subtitle": "Contract liability activity from September through December 2025",
      "entity_header": "Customer",
      "term_header": "Contract Months",
      "notes_header": "Notes",
      "account_header": "Liability Account",
      "totals_label": "Period Totals",
      "gl_key": "services_2355",
      "months": [
        {
          "slug": "sep",
          "label": "Sep",
          "adds_subheader": "Billings",
          "release_subheader": "Revenue"
        },
        {
          "slug": "oct",
          "label": "Oct",
          "adds_subheader": "Billings",
          "release_subheader": "Revenue"
        },
        {
          "slug": "nov",
          "label": "Nov",
          "adds_subheader": "Billings",
          "release_subheader": "Revenue"
        },
        {
          "slug": "dec",
          "label": "Dec",
          "adds_subheader": "Billings",
          "release_subheader": "Revenue"
        }
      ],
      "summary_labels": {
        "section": "Services Contract Liability (Acct 2355)",
        "total": "Services billings in period",
        "ending": "Services revenue recognized in period",
        "gl": "Services GL balance at 12/31/2025"
      }
    }
  ]
}
EXPECTED_GL = {
  "subscriptions_2350": {
    "sep": 21000.0,
    "oct": 26000.0,
    "nov": 20000.0,
    "dec": 11000.0
  },
  "services_2355": {
    "sep": 12000.0,
    "oct": 14750.0,
    "nov": 10000.0,
    "dec": 6250.0
  }
}
EXPECTED_ROWS = {
  "Subscriptions #2350": [
    {
      "entity": "Aurora Transit",
      "beginning_balance": 0.0,
      "term_months": 12,
      "comments": "Enterprise transit subscription",
      "account_number": 2350,
      "sep_adds": 20000,
      "sep_release": 5000,
      "sep_ending_balance": 15000.0,
      "oct_adds": 0,
      "oct_release": 5000,
      "oct_ending_balance": 10000.0,
      "nov_adds": 5000,
      "nov_release": 5000,
      "nov_ending_balance": 10000.0,
      "dec_adds": 0,
      "dec_release": 5000,
      "dec_ending_balance": 5000.0
    },
    {
      "entity": "Beacon Fleet",
      "beginning_balance": 0.0,
      "term_months": 8,
      "comments": "Fleet onboarding subscription",
      "account_number": 2350,
      "sep_adds": 0,
      "sep_release": 0,
      "sep_ending_balance": 0.0,
      "oct_adds": 16000,
      "oct_release": 4000,
      "oct_ending_balance": 12000.0,
      "nov_adds": 0,
      "nov_release": 4000,
      "nov_ending_balance": 8000.0,
      "dec_adds": 0,
      "dec_release": 4000,
      "dec_ending_balance": 4000.0
    },
    {
      "entity": "Comet Riders",
      "beginning_balance": 0.0,
      "term_months": 6,
      "comments": "Riders annual renewal",
      "account_number": 2350,
      "sep_adds": 8000,
      "sep_release": 2000,
      "sep_ending_balance": 6000.0,
      "oct_adds": 0,
      "oct_release": 2000,
      "oct_ending_balance": 4000.0,
      "nov_adds": 0,
      "nov_release": 2000,
      "nov_ending_balance": 2000.0,
      "dec_adds": 2000,
      "dec_release": 2000,
      "dec_ending_balance": 2000.0
    }
  ],
  "Services #2355": [
    {
      "entity": "Delta Install",
      "beginning_balance": 0.0,
      "term_months": 4,
      "comments": "Implementation package",
      "account_number": 2355,
      "sep_adds": 12000,
      "sep_release": 3000,
      "sep_ending_balance": 9000.0,
      "oct_adds": 0,
      "oct_release": 3000,
      "oct_ending_balance": 6000.0,
      "nov_adds": 0,
      "nov_release": 3000,
      "nov_ending_balance": 3000.0,
      "dec_adds": 0,
      "dec_release": 3000,
      "dec_ending_balance": 0.0
    },
    {
      "entity": "Ember Training",
      "beginning_balance": 0.0,
      "term_months": 6,
      "comments": "Training services reserve",
      "account_number": 2355,
      "sep_adds": 0,
      "sep_release": 0,
      "sep_ending_balance": 0.0,
      "oct_adds": 9000,
      "oct_release": 2250,
      "oct_ending_balance": 6750.0,
      "nov_adds": 0,
      "nov_release": 2250,
      "nov_ending_balance": 4500.0,
      "dec_adds": 3000,
      "dec_release": 2250,
      "dec_ending_balance": 5250.0
    },
    {
      "entity": "Focal Support",
      "beginning_balance": 0.0,
      "term_months": 6,
      "comments": "Premium support bundle",
      "account_number": 2355,
      "sep_adds": 4000,
      "sep_release": 1000,
      "sep_ending_balance": 3000.0,
      "oct_adds": 0,
      "oct_release": 1000,
      "oct_ending_balance": 2000.0,
      "nov_adds": 2000,
      "nov_release": 1500,
      "nov_ending_balance": 2500.0,
      "dec_adds": 0,
      "dec_release": 1500,
      "dec_ending_balance": 1000.0
    }
  ]
}
MONTHS = ["sep", "oct", "nov", "dec"]
OUTPUT_FILENAME = 'Meridian_Contract_Liability_12-25.xlsx'
EXPECTED_SHEETS = SPEC['sheet_order']

def fail(message: str) -> None:
    raise SystemExit(f'FAIL: {message}')

def ok(message: str) -> None:
    print(f'PASS: {message}')

def assert_true(condition: bool, message: str) -> None:
    if not condition:
        fail(message)

def near(a, b, tolerance=0.01) -> bool:
    return abs(float(a or 0) - float(b or 0)) <= tolerance

def text(value) -> str:
    return '' if value is None else str(value).strip()

def normalize_formula(value) -> str:
    raw = text(value)
    if raw.startswith('='):
        raw = raw[1:]
    return raw.replace('$', '').replace(' ', '').upper()

def col_letter(index: int) -> str:
    result = ''
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result

task_dir = Path(__file__).resolve().parent.parent
output_path = Path('/root') / OUTPUT_FILENAME if (Path('/root') / OUTPUT_FILENAME).exists() else task_dir / OUTPUT_FILENAME
assert_true(output_path.exists(), f'Workbook not found: {output_path}')
ok('Output workbook exists')

workbook = load_workbook(output_path, data_only=False)
assert_true(set(workbook.sheetnames) == set(EXPECTED_SHEETS), f'Sheet names must be {EXPECTED_SHEETS}')
ok('Sheet names and order are correct')

def validate_detail(detail: dict) -> tuple[int, int, int, str]:
    sheet_name = detail['sheet_name']
    rows = EXPECTED_ROWS[sheet_name]
    ws = workbook[sheet_name]
    assert_true(text(ws['A1'].value) == detail['company'], f'{sheet_name}!A1 company mismatch')
    assert_true(text(ws['A2'].value) == detail['title'], f'{sheet_name}!A2 title mismatch')

    start_row = 6
    end_row = start_row + len(rows) - 1
    totals_row = end_row + 1
    ending_row = totals_row + 1
    variance_row = ending_row + 1
    gl_row = variance_row + 1

    term_col_idx = 2 + len(MONTHS) * 3 + 1
    notes_col_idx = term_col_idx + 1
    account_col_idx = term_col_idx + 2
    term_col = col_letter(term_col_idx)
    last_end_col = col_letter(2 + len(MONTHS) * 3)

    for offset, row in enumerate(rows):
        excel_row = start_row + offset
        assert_true(text(ws[f'A{excel_row}'].value) == row['entity'], f'{sheet_name} entity mismatch on row {excel_row}')
        assert_true(near(ws[f'B{excel_row}'].value, row['beginning_balance']), f'{sheet_name} beginning balance mismatch on row {excel_row}')
        for index, slug in enumerate(MONTHS):
            base_col = 3 + index * 3
            add_col = col_letter(base_col)
            rel_col = col_letter(base_col + 1)
            end_col = col_letter(base_col + 2)
            assert_true(near(ws[f'{add_col}{excel_row}'].value, row[f'{slug}_adds']), f'{sheet_name} {add_col}{excel_row} mismatch')
            assert_true(near(ws[f'{rel_col}{excel_row}'].value, row[f'{slug}_release']), f'{sheet_name} {rel_col}{excel_row} mismatch')
            assert_true(near(ws[f'{end_col}{excel_row}'].value, row[f'{slug}_ending_balance']), f'{sheet_name} {end_col}{excel_row} mismatch')
        assert_true(near(ws[f'{term_col}{excel_row}'].value, row['term_months']), f'{sheet_name} term months mismatch on row {excel_row}')
        notes_col = col_letter(notes_col_idx)
        account_col = col_letter(account_col_idx)
        assert_true(text(ws[f'{notes_col}{excel_row}'].value) == row['comments'], f'{sheet_name} comments mismatch on row {excel_row}')
        assert_true(near(ws[f'{account_col}{excel_row}'].value, row['account_number']), f'{sheet_name} account mismatch on row {excel_row}')
    ok(f'{sheet_name} row-level data matches expectation')

    assert_true(text(ws[f'A{totals_row}'].value).lower() == 'period totals', f'{sheet_name} totals row label mismatch')
    for col_idx in range(2, 2 + len(MONTHS) * 3 + 1):
        col = col_letter(col_idx)
        expected_formula = f'SUM({col}{start_row}:{col}{end_row})'
        assert_true(normalize_formula(ws[f'{col}{totals_row}'].value) == normalize_formula(expected_formula), f'{sheet_name} {col}{totals_row} totals formula mismatch')
    add_cols = [col_letter(3 + index * 3) for index in range(len(MONTHS))]
    rel_cols = [col_letter(4 + index * 3) for index in range(len(MONTHS))]
    expected_total_formula = '+'.join(f'{col}{totals_row}' for col in add_cols)
    assert_true(normalize_formula(ws[f'{term_col}{totals_row}'].value) == normalize_formula(expected_total_formula), f'{sheet_name} totals term formula mismatch')

    assert_true(text(ws[f'A{ending_row}'].value).lower() == 'ending balance', f'{sheet_name} ending row label mismatch')
    first_end = col_letter(5)
    assert_true(normalize_formula(ws[f'{first_end}{ending_row}'].value) == normalize_formula(f'B{totals_row}+C{totals_row}-D{totals_row}'), f'{sheet_name} first ending formula mismatch')
    for index in range(1, len(MONTHS)):
        prev_end = col_letter(5 + (index - 1) * 3)
        add_col = col_letter(3 + index * 3)
        rel_col = col_letter(4 + index * 3)
        end_col = col_letter(5 + index * 3)
        expected_formula = f'{prev_end}{totals_row}+{add_col}{totals_row}-{rel_col}{totals_row}'
        assert_true(normalize_formula(ws[f'{end_col}{ending_row}'].value) == normalize_formula(expected_formula), f'{sheet_name} {end_col}{ending_row} formula mismatch')
    release_formula = '+'.join(f'{col}{totals_row}' for col in rel_cols)
    assert_true(normalize_formula(ws[f'{term_col}{ending_row}'].value) == normalize_formula(release_formula), f'{sheet_name} ending term formula mismatch')

    assert_true(text(ws[f'A{variance_row}'].value).lower() == 'variance', f'{sheet_name} variance row label mismatch')
    assert_true(normalize_formula(ws[f'{term_col}{variance_row}'].value) == normalize_formula(f'{term_col}{gl_row}-{last_end_col}{gl_row}'), f'{sheet_name} variance formula mismatch')

    assert_true(text(ws[f'A{gl_row}'].value).lower() == 'gl balance', f'{sheet_name} GL row label mismatch')
    for index, slug in enumerate(MONTHS):
        end_col = col_letter(5 + index * 3)
        assert_true(near(ws[f'{end_col}{gl_row}'].value, EXPECTED_GL[detail['gl_key']][slug]), f'{sheet_name} GL value mismatch for {slug}')
    assert_true(normalize_formula(ws[f'{term_col}{gl_row}'].value) == normalize_formula(f'{term_col}{totals_row}-{term_col}{ending_row}'), f'{sheet_name} GL term formula mismatch')
    ok(f'{sheet_name} control rows and GL linkage are correct')
    return totals_row, ending_row, gl_row, term_col

first_info = validate_detail(SPEC['details'][0])
second_info = validate_detail(SPEC['details'][1])

summary = workbook[SPEC['summary']['sheet_name']]
assert_true(text(summary['A1'].value) == SPEC['summary']['company'], 'Summary company mismatch')
assert_true(text(summary['A2'].value) == SPEC['summary']['title'], 'Summary title mismatch')
assert_true(text(summary['A3'].value) == SPEC['summary']['period_text'], 'Summary period mismatch')
assert_true(normalize_formula(summary['B7'].value) == normalize_formula("'Subscriptions #2350'!" + f"{first_info[3]}{first_info[0]}"), 'Summary B7 mismatch')
assert_true(normalize_formula(summary['B8'].value) == normalize_formula("'Subscriptions #2350'!" + f"{first_info[3]}{first_info[1]}"), 'Summary B8 mismatch')
assert_true(normalize_formula(summary['B9'].value) == normalize_formula("'Subscriptions #2350'!" + f"{first_info[3]}{first_info[2]}"), 'Summary B9 mismatch')
assert_true(normalize_formula(summary['B12'].value) == normalize_formula("'Services #2355'!" + f"{second_info[3]}{second_info[0]}"), 'Summary B12 mismatch')
assert_true(normalize_formula(summary['B13'].value) == normalize_formula("'Services #2355'!" + f"{second_info[3]}{second_info[1]}"), 'Summary B13 mismatch')
assert_true(normalize_formula(summary['B14'].value) == normalize_formula("'Services #2355'!" + f"{second_info[3]}{second_info[2]}"), 'Summary B14 mismatch')
assert_true(normalize_formula(summary['B16'].value) == normalize_formula('B9+B14'), 'Summary B16 mismatch')
ok('Summary links are correct')



for sheet_name in EXPECTED_SHEETS[1:]:
    ws = workbook[sheet_name]
    assert_true(text(ws['A40'].value) == '', f'{sheet_name}!A40 should be blank after stale rows are cleared')
    assert_true(text(ws['A6'].value) != 'STALE ROW - REMOVE', f'{sheet_name}!A6 still contains template placeholder text')
ok('Template stale rows were cleared')

print('All tests passed.')
