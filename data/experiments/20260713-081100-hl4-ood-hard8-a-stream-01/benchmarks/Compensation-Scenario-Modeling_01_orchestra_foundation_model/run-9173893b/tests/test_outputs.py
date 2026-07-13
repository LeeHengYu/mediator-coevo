from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

OUTPUT_FILE = Path("/root/Orchestra_Compensation.xlsx")
SOURCE_FILE = Path("/root/orchestra_assumptions_and_roster.xlsx")


def norm_formula(value):
    if value is None:
        return None
    return str(value).replace(" ", "")


@pytest.fixture(scope="module")
def output_wb():
    assert OUTPUT_FILE.exists(), f"Missing required output file: {OUTPUT_FILE}"
    return load_workbook(OUTPUT_FILE, data_only=False)


@pytest.fixture(scope="module")
def source_wb():
    assert SOURCE_FILE.exists(), f"Missing required source file: {SOURCE_FILE}"
    return load_workbook(SOURCE_FILE, data_only=False)


def test_sheet_names(output_wb):
    assert output_wb.sheetnames == [
        "Summary",
        "Assumptions",
        "Roster",
        "Calculations --->",
        "EE Calcs (Current)",
        "EE Calcs (Yr+1)",
        "EE Calcs (Yr+2)",
    ]


def test_defined_names(output_wb):
    expected = {
        "MWS_Yr1": "Summary!$C$5",
        "MWS_Yr2": "Summary!$D$5",
        "MWS_Yr3": "Summary!$E$5",
        "Princ_Yr1": "Summary!$C$7",
        "Princ_Yr2": "Summary!$D$7",
        "Princ_Yr3": "Summary!$E$7",
        "AssocPr_Yr1": "Summary!$C$8",
        "AssocPr_Yr2": "Summary!$D$8",
        "AssocPr_Yr3": "Summary!$E$8",
        "AsstPr_Yr1": "Summary!$C$9",
        "AsstPr_Yr2": "Summary!$D$9",
        "AsstPr_Yr3": "Summary!$E$9",
        "Media_Yr1": "Summary!$C$10",
        "Media_Yr2": "Summary!$D$10",
        "Media_Yr3": "Summary!$E$10",
        "WHLim_Yr1": "Summary!$C$12",
        "WHLim_Yr2": "Summary!$D$12",
        "WHLim_Yr3": "Summary!$E$12",
        "WHto7k_Yr1": "Summary!$C$13",
        "WHto7k_Yr2": "Summary!$D$13",
        "WHto7k_Yr3": "Summary!$E$13",
        "WH7ktoLim_Yr1": "Summary!$C$14",
        "WH7ktoLim_Yr2": "Summary!$D$14",
        "WH7ktoLim_Yr3": "Summary!$E$14",
        "WHoverLim_Yr1": "Summary!$C$15",
        "WHoverLim_Yr2": "Summary!$D$15",
        "WHoverLim_Yr3": "Summary!$E$15",
        "Sr5to9_Yr1": "Summary!$C$17",
        "Sr5to9_Yr2": "Summary!$D$17",
        "Sr5to9_Yr3": "Summary!$E$17",
        "Sr10to14_Yr1": "Summary!$C$18",
        "Sr10to14_Yr2": "Summary!$D$18",
        "Sr10to14_Yr3": "Summary!$E$18",
        "Sr15to19_Yr1": "Summary!$C$19",
        "Sr15to19_Yr2": "Summary!$D$19",
        "Sr15to19_Yr3": "Summary!$E$19",
        "Sr20to24_Yr1": "Summary!$C$20",
        "Sr20to24_Yr2": "Summary!$D$20",
        "Sr20to24_Yr3": "Summary!$E$20",
        "Sr25up_Yr1": "Summary!$C$21",
        "Sr25up_Yr2": "Summary!$D$21",
        "Sr25up_Yr3": "Summary!$E$21",
    }

    for name, expected_ref in expected.items():
        assert name in output_wb.defined_names, f"Missing defined name: {name}"
        actual = output_wb.defined_names[name].attr_text
        assert actual == expected_ref, f"Defined name {name} expected {expected_ref}, got {actual}"


def test_summary_structure_and_formulas(output_wb):
    ws = output_wb["Summary"]

    assert ws["B1"].value == "Renaissance Popular Orchestra - Musician Compensation"
    assert ws["C5"].value == 2395
    assert norm_formula(ws["D5"].value) == "=C5*(1+D6)"
    assert norm_formula(ws["E5"].value) == "=D5*(1+E6)"
    assert ws["D6"].value == 0.05
    assert ws["E6"].value == 0.05
    assert norm_formula(ws["D12"].value) == "=WHLim_Yr1*1.03"
    assert norm_formula(ws["E12"].value) == "=D12*1.03"

    expected_labels = [
        "MWS",
        "Overscale",
        "Principal Pay",
        "Media Exploitation",
        "Seniority",
        "Payroll Tax",
    ]
    for i, label in enumerate(expected_labels, start=26):
        assert ws[f"B{i}"].value == label

    calc_col_groups = {
        26: ["H", "I", "J", "K"],
        27: ["L", "M", "N", "O"],
        28: ["P", "Q", "R", "S"],
        29: ["T", "U", "V", "W"],
        30: ["X", "Y", "Z", "AA"],
        31: ["AF", "AG", "AH", "AI"],
    }

    year_sheets = ["EE Calcs (Current)", "EE Calcs (Yr+1)", "EE Calcs (Yr+2)"]
    summary_start_cols = [3, 7, 11]  # C, G, K

    for row_num, calc_cols in calc_col_groups.items():
        for sheet_name, start_col in zip(year_sheets, summary_start_cols):
            for offset, calc_col in enumerate(calc_cols):
                summary_col = get_column_letter(start_col + offset)
                expected_formula = f"='{sheet_name}'!{calc_col}107"
                assert norm_formula(ws[f"{summary_col}{row_num}"].value) == norm_formula(expected_formula)

    for col_idx in range(3, 15):  # C:N
        col = get_column_letter(col_idx)
        assert norm_formula(ws[f"{col}32"].value) == f"=SUM({col}26:{col}31)"

    expected_yoy = {
        "G33": "=G32/C32-1",
        "H33": "=H32/D32-1",
        "I33": "=I32/E32-1",
        "J33": "=J32/F32-1",
        "K33": "=K32/G32-1",
        "L33": "=L32/H32-1",
        "M33": "=M32/I32-1",
        "N33": "=N32/J32-1",
    }
    for cell, formula in expected_yoy.items():
        assert norm_formula(ws[cell].value) == formula


def test_roster_and_assumptions_migration(output_wb, source_wb):
    src_assump = source_wb["Assumptions"]
    out_assump = output_wb["Assumptions"]

    assert "Renaissance Popular Orchestra" in str(out_assump["B1"].value)
    assert out_assump["B3"].value == src_assump["B3"].value
    assert "Payroll Tax" in str(out_assump["B18"].value)

    for row in range(13, 18):
        assert out_assump[f"C{row}"].value == src_assump[f"C{row}"].value
        assert out_assump[f"D{row}"].value == src_assump[f"D{row}"].value

    src_roster = source_wb["Roster"]
    out_roster = output_wb["Roster"]

    # Source data is B6:F108. Output is A4:E106.
    for i, src_row in enumerate(range(6, 109), start=4):
        src_vals = [src_roster.cell(src_row, c).value for c in range(2, 7)]
        out_vals = [out_roster.cell(i, c).value for c in range(1, 6)]
        assert out_vals == src_vals, f"Roster row mismatch at output row {i}"


@pytest.mark.parametrize(
    "sheet_name,year_token,service_ref",
    [
        ("EE Calcs (Current)", "Yr1", None),
        ("EE Calcs (Yr+1)", "Yr2", "'EE Calcs (Current)'"),
        ("EE Calcs (Yr+2)", "Yr3", "'EE Calcs (Yr+1)'"),
    ],
)
def test_calc_sheet_formulas(output_wb, sheet_name, year_token, service_ref):
    ws = output_wb[sheet_name]

    assert ws["A3"].value == "Employee #"
    assert ws["B3"].value == "Titled Positions"
    assert ws["C3"].value == "Overscale"
    assert ws["D3"].value == "Years of Service"
    assert ws["E3"].value == "Notes"
    assert ws["F3"].value == "Principal"
    assert ws["G3"].value == "Seniority"

    if service_ref is None:
        assert isinstance(ws["D4"].value, (int, float))
    else:
        d4 = norm_formula(ws["D4"].value)
        assert d4 is not None
        assert norm_formula(service_ref) in d4
        assert "+1" in d4

    assert f"MWS_{year_token}" in str(ws["F4"].value)
    assert f"Sr5to9_{year_token}" in str(ws["G4"].value)
    assert norm_formula(ws["H4"].value) == f"=MWS_{year_token}*13"
    assert norm_formula(ws["L4"].value) == "=IF(ISBLANK($E4),$C4*13,($C4-$F4)*13)"
    assert norm_formula(ws["P4"].value) == "=$F4*13"
    assert norm_formula(ws["T4"].value) == f"=Media_{year_token}*13"
    assert norm_formula(ws["AB4"].value) == "=SUMIFS($H4:$AA4,$H$3:$AA$3,AB$3)"
    assert f"WHto7k_{year_token}" in str(ws["AF4"].value)
    assert f"WHLim_{year_token}" in str(ws["AF4"].value)
    assert f"WHoverLim_{year_token}" in str(ws["AF4"].value)

    for col_idx in range(8, 36):  # H..AI
        col = get_column_letter(col_idx)
        expected = f"=SUM({col}4:{col}106)"
        assert norm_formula(ws[f"{col}107"].value) == expected
