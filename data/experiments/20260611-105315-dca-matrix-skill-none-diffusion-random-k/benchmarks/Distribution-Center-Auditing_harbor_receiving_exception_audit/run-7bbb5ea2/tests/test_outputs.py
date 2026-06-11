"""Verification tests for harbor_receiving_exception_audit."""

from collections import defaultdict
from pathlib import Path
import os

import pytest
from docx import Document
from openpyxl import load_workbook

ROOT_DIR = os.environ.get('TASK_ROOT', '/root')
SOURCE_XLSX = f"{ROOT_DIR}/Receiving_Log.xlsx"
OUTPUT_XLSX = f"{ROOT_DIR}/Receiving_Exception_Audit.xlsx"
OUTPUT_DOCX = f"{ROOT_DIR}/Receiving_Exception_Brief.docx"

BASE_HEADERS = [
    'Receipt ID',
    'Item Code',
    'Expected Qty',
    'Received Qty',
    'Storage Class',
    'Temp Status',
    'Supplier',
    'Dock',
]

FORMATTED_HEADERS = BASE_HEADERS + [
    'Qty Variance',
    'Cold Chain Error',
    'Total Errors',
    'Error Summary',
]

SUMMARY_HEADERS = [
    'Item Code',
    'Supplier',
    'Qty Variance Errors',
    'Cold Chain Errors',
    'Total Errors',
]

def calc_flags(row):
    qty_variance = 1 if row['Received Qty'] != row['Expected Qty'] else 0
    storage_class = str(row['Storage Class'] or '').strip().upper()
    temp_status = str(row['Temp Status'] or '').strip().upper()
    cold_chain = 1 if storage_class in {'CHILLED', 'FROZEN'} and temp_status != 'OK' else 0
    total = qty_variance + cold_chain
    if total == 0:
        summary = 'None'
    elif qty_variance and cold_chain:
        summary = 'Qty Variance, Cold Chain Error'
    elif qty_variance:
        summary = 'Qty Variance'
    else:
        summary = 'Cold Chain Error'
    return qty_variance, cold_chain, total, summary

@pytest.fixture(scope='module')
def source_rows():
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb['Receipts']
    headers = [ws.cell(1, c).value for c in range(1, 9)]
    assert headers == BASE_HEADERS
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 9)]
        if all(v is None for v in vals):
            continue
        rows.append(dict(zip(BASE_HEADERS, vals)))
    assert rows
    return rows

@pytest.fixture(scope='module')
def output_wb():
    assert Path(OUTPUT_XLSX).exists(), f'Missing output workbook: {OUTPUT_XLSX}'
    return load_workbook(OUTPUT_XLSX, data_only=True)

def expected_summary_rows(source_rows):
    agg = defaultdict(lambda: [0, 0, 0])
    total_qty = total_cold = total_errors = 0
    item_totals = defaultdict(int)
    for row in source_rows:
        qty, cold, total, _ = calc_flags(row)
        key = (str(row['Item Code']), str(row['Supplier']))
        agg[key][0] += qty
        agg[key][1] += cold
        agg[key][2] += total
        total_qty += qty
        total_cold += cold
        total_errors += total
        item_totals[str(row['Item Code'])] += total
    rows = []
    for (item_code, supplier), values in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1])):
        if values[2] > 0:
            rows.append((item_code, supplier, values[0], values[1], values[2]))
    return rows, (total_qty, total_cold, total_errors), item_totals

def test_required_sheets_exist(output_wb):
    assert {'RawData', 'Formatted Data', 'Summary'}.issubset(set(output_wb.sheetnames))

def test_rawdata_copies_source_exactly(output_wb, source_rows):
    ws = output_wb['RawData']
    headers = [ws.cell(1, c).value for c in range(1, 9)]
    assert headers == BASE_HEADERS
    actual = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 9)]
        if all(v is None for v in vals):
            continue
        actual.append(vals)
    expected = [[row[h] for h in BASE_HEADERS] for row in source_rows]
    assert actual == expected

def test_formatted_data_logic(output_wb, source_rows):
    ws = output_wb['Formatted Data']
    headers = [ws.cell(1, c).value for c in range(1, 13)]
    assert headers == FORMATTED_HEADERS
    for i, source_row in enumerate(source_rows, start=2):
        out = [ws.cell(i, c).value for c in range(1, 13)]
        assert out[:8] == [source_row[h] for h in BASE_HEADERS]
        expected_qty, expected_cold, expected_total, expected_summary = calc_flags(source_row)
        assert int(out[8]) == expected_qty
        assert int(out[9]) == expected_cold
        assert int(out[10]) == expected_total
        assert str(out[11]).strip() == expected_summary

def test_summary_sheet_matches_expected(output_wb, source_rows):
    ws = output_wb['Summary']
    headers = [ws.cell(1, c).value for c in range(1, 6)]
    assert headers == SUMMARY_HEADERS
    expected_rows, expected_totals, _ = expected_summary_rows(source_rows)
    actual_rows = []
    grand_total = None
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 6)]
        if all(v is None for v in vals):
            continue
        if str(vals[0]).strip() == 'Grand Total':
            grand_total = (str(vals[0]), str(vals[1]), int(vals[2]), int(vals[3]), int(vals[4]))
        else:
            actual_rows.append((str(vals[0]), str(vals[1]), int(vals[2]), int(vals[3]), int(vals[4])))
    assert actual_rows == expected_rows
    assert grand_total == ('Grand Total', '-', expected_totals[0], expected_totals[1], expected_totals[2])

def test_word_summary_content(source_rows):
    assert Path(OUTPUT_DOCX).exists(), f'Missing output docx: {OUTPUT_DOCX}'
    _, expected_totals, item_totals = expected_summary_rows(source_rows)
    doc = Document(OUTPUT_DOCX)
    text = '\n'.join(p.text for p in doc.paragraphs if p.text.strip()).lower()
    assert 'qty variance' in text
    assert 'cold chain' in text
    for value in expected_totals:
        assert str(value) in text
    top_items = [item for item, total in sorted(item_totals.items(), key=lambda x: (-x[1], x[0])) if total > 0][:4]
    mentioned = sum(1 for item in top_items if item.lower() in text)
    assert mentioned >= 2, 'Word summary must mention at least two high-priority item codes'
    assert any(marker in text for marker in ['recommend', 'review', 'should', 'action', 'prioritize'])
