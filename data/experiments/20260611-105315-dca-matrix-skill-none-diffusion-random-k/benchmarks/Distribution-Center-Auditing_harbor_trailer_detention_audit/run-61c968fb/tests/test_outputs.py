"""Verification tests for harbor_trailer_detention_audit."""

from collections import defaultdict
from pathlib import Path
import os

import pytest
from docx import Document
from openpyxl import load_workbook

ROOT_DIR = os.environ.get('TASK_ROOT', '/root')
SOURCE_XLSX = f"{ROOT_DIR}/Trailer_Detention_Log.xlsx"
OUTPUT_XLSX = f"{ROOT_DIR}/Trailer_Detention_Audit.xlsx"
OUTPUT_DOCX = f"{ROOT_DIR}/Trailer_Detention_Brief.docx"

BASE_HEADERS = [
    'Load ID',
    'Carrier',
    'Allowed Hold Hours',
    'Actual Hold Hours',
    'Seal Required',
    'Seal Status',
    'Yard',
    'Dispatcher',
]

FORMATTED_HEADERS = BASE_HEADERS + ['Detention Overrun', 'Seal Error', 'Total Errors', 'Error Summary']
SUMMARY_HEADERS = ['Carrier', 'Yard', 'Detention Overrun Errors', 'Seal Errors', 'Total Errors']


def calc_flags(row):
    detention = 1 if float(row['Actual Hold Hours']) > float(row['Allowed Hold Hours']) else 0
    seal_required = str(row['Seal Required'] or '').strip().upper()
    seal_status = str(row['Seal Status'] or '').strip().upper()
    seal_error = 1 if seal_required == 'YES' and seal_status != 'VERIFIED' else 0
    total = detention + seal_error
    if total == 0:
        summary = 'None'
    elif detention and seal_error:
        summary = 'Detention Overrun, Seal Error'
    elif detention:
        summary = 'Detention Overrun'
    else:
        summary = 'Seal Error'
    return detention, seal_error, total, summary


@pytest.fixture(scope='module')
def source_rows():
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb['Detention']
    headers = [ws.cell(1, c).value for c in range(1, 9)]
    assert headers == BASE_HEADERS
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 9)]
        if all(v is None for v in vals):
            continue
        rows.append(dict(zip(BASE_HEADERS, vals)))
    assert rows
    return rows


@pytest.fixture(scope='module')
def output_wb():
    assert Path(OUTPUT_XLSX).exists(), f'Missing output workbook: {OUTPUT_XLSX}'
    return load_workbook(OUTPUT_XLSX, data_only=True)


def expected_summary_rows(source_rows):
    agg = defaultdict(lambda: [0, 0, 0])
    carrier_totals = defaultdict(int)
    total_detention = total_seal = total_errors = 0
    for row in source_rows:
        detention, seal, total, _ = calc_flags(row)
        key = (str(row['Carrier']), str(row['Yard']))
        agg[key][0] += detention
        agg[key][1] += seal
        agg[key][2] += total
        carrier_totals[str(row['Carrier'])] += total
        total_detention += detention
        total_seal += seal
        total_errors += total
    rows = []
    for (carrier, yard), values in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1])):
        if values[2] > 0:
            rows.append((carrier, yard, values[0], values[1], values[2]))
    return rows, (total_detention, total_seal, total_errors), carrier_totals


def test_required_sheets_exist(output_wb):
    assert {'RawData', 'Formatted Data', 'Summary'}.issubset(set(output_wb.sheetnames))


def test_rawdata_copies_source_exactly(output_wb, source_rows):
    ws = output_wb['RawData']
    headers = [ws.cell(1, c).value for c in range(1, 9)]
    assert headers == BASE_HEADERS
    actual = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 9)]
        if all(v is None for v in vals):
            continue
        actual.append(vals)
    expected = [[row[h] for h in BASE_HEADERS] for row in source_rows]
    assert actual == expected


def test_formatted_data_logic(output_wb, source_rows):
    ws = output_wb['Formatted Data']
    headers = [ws.cell(1, c).value for c in range(1, 13)]
    assert headers == FORMATTED_HEADERS
    for i, row in enumerate(source_rows, start=2):
        out = [ws.cell(i, c).value for c in range(1, 13)]
        assert out[:8] == [row[h] for h in BASE_HEADERS]
        expected_det, expected_seal, expected_total, expected_summary = calc_flags(row)
        assert int(out[8]) == expected_det
        assert int(out[9]) == expected_seal
        assert int(out[10]) == expected_total
        assert str(out[11]).strip() == expected_summary


def test_summary_sheet_matches_expected(output_wb, source_rows):
    ws = output_wb['Summary']
    headers = [ws.cell(1, c).value for c in range(1, 6)]
    assert headers == SUMMARY_HEADERS
    expected_rows, expected_totals, _ = expected_summary_rows(source_rows)
    actual_rows = []
    grand_total = None
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 6)]
        if all(v is None for v in vals):
            continue
        if str(vals[0]).strip() == 'Grand Total':
            grand_total = (str(vals[0]), str(vals[1]), int(vals[2]), int(vals[3]), int(vals[4]))
        else:
            actual_rows.append((str(vals[0]), str(vals[1]), int(vals[2]), int(vals[3]), int(vals[4])))
    assert actual_rows == expected_rows
    assert grand_total == ('Grand Total', '-', expected_totals[0], expected_totals[1], expected_totals[2])


def test_word_summary_content(source_rows):
    assert Path(OUTPUT_DOCX).exists(), f'Missing output docx: {OUTPUT_DOCX}'
    _, expected_totals, carrier_totals = expected_summary_rows(source_rows)
    doc = Document(OUTPUT_DOCX)
    text = '\n'.join(p.text for p in doc.paragraphs if p.text.strip()).lower()
    assert 'detention overrun' in text
    assert 'seal error' in text
    for value in expected_totals:
        assert str(value) in text
    top_carriers = [c for c, total in sorted(carrier_totals.items(), key=lambda x: (-x[1], x[0])) if total > 0][:4]
    mentioned = sum(1 for c in top_carriers if c.lower() in text)
    assert mentioned >= 2, 'Word summary must mention at least two high-priority carriers'
    assert any(marker in text for marker in ['recommend', 'review', 'should', 'action', 'prioritize'])
