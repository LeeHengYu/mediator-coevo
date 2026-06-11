"""Verification tests for harbor_promo_register_audit."""

from collections import defaultdict
from pathlib import Path
import datetime as dt
import os

import pytest
from docx import Document
from openpyxl import load_workbook

ROOT_DIR = os.environ.get('TASK_ROOT', '/root')
SOURCE_XLSX = f"{ROOT_DIR}/Promo_Price_Check_Source.xlsx"
OUTPUT_XLSX = f"{ROOT_DIR}/Promo_Register_Audit.xlsx"
OUTPUT_DOCX = f"{ROOT_DIR}/Promo_Register_Brief.docx"

BASE_HEADERS = [
    'Promo ID',
    'SKU',
    'Promo Price',
    'Register Price',
    'Promo Start Date',
    'Sale Date',
    'Promo End Date',
    'Store ID',
]

FORMATTED_HEADERS = BASE_HEADERS + ['Price Error', 'Window Error', 'Total Errors', 'Error Summary']
SUMMARY_HEADERS = ['SKU', 'Store ID', 'Price Errors', 'Window Errors', 'Total Errors']

def parse_date(value):
    if hasattr(value, 'date'):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return dt.date.fromisoformat(str(value))

def calc_flags(row):
    price_error = 1 if row['Register Price'] != row['Promo Price'] else 0
    start_date = parse_date(row['Promo Start Date'])
    sale_date = parse_date(row['Sale Date'])
    end_date = parse_date(row['Promo End Date'])
    window_error = 1 if sale_date < start_date or sale_date > end_date else 0
    total = price_error + window_error
    if total == 0:
        summary = 'None'
    elif price_error and window_error:
        summary = 'Price Error, Window Error'
    elif price_error:
        summary = 'Price Error'
    else:
        summary = 'Window Error'
    return price_error, window_error, total, summary

@pytest.fixture(scope='module')
def source_rows():
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb['SalesAudit']
    headers = [ws.cell(1, c).value for c in range(1, 9)]
    assert headers == BASE_HEADERS
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 9)]
        if all(v is None for v in vals):
            continue
        rows.append(dict(zip(BASE_HEADERS, vals)))
    assert rows
    return rows

@pytest.fixture(scope='module')
def output_wb():
    assert Path(OUTPUT_XLSX).exists(), f'Missing output workbook: {OUTPUT_XLSX}'
    return load_workbook(OUTPUT_XLSX, data_only=True)

def expected_summary_rows(source_rows):
    agg = defaultdict(lambda: [0, 0, 0])
    total_price = total_window = total_errors = 0
    sku_totals = defaultdict(int)
    for row in source_rows:
        price, window, total, _ = calc_flags(row)
        key = (str(row['SKU']), str(row['Store ID']))
        agg[key][0] += price
        agg[key][1] += window
        agg[key][2] += total
        total_price += price
        total_window += window
        total_errors += total
        sku_totals[str(row['SKU'])] += total
    rows = []
    for (sku, store_id), values in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1])):
        if values[2] > 0:
            rows.append((sku, store_id, values[0], values[1], values[2]))
    return rows, (total_price, total_window, total_errors), sku_totals

def test_required_sheets_exist(output_wb):
    assert {'RawData', 'Formatted Data', 'Summary'}.issubset(set(output_wb.sheetnames))

def test_rawdata_copies_source_exactly(output_wb, source_rows):
    ws = output_wb['RawData']
    headers = [ws.cell(1, c).value for c in range(1, 9)]
    assert headers == BASE_HEADERS
    actual = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 9)]
        if all(v is None for v in vals):
            continue
        actual.append(vals)
    expected = [[row[h] for h in BASE_HEADERS] for row in source_rows]
    assert actual == expected

def test_formatted_data_logic(output_wb, source_rows):
    ws = output_wb['Formatted Data']
    headers = [ws.cell(1, c).value for c in range(1, 13)]
    assert headers == FORMATTED_HEADERS
    for i, source_row in enumerate(source_rows, start=2):
        out = [ws.cell(i, c).value for c in range(1, 13)]
        assert out[:8] == [source_row[h] for h in BASE_HEADERS]
        expected_price, expected_window, expected_total, expected_summary = calc_flags(source_row)
        assert int(out[8]) == expected_price
        assert int(out[9]) == expected_window
        assert int(out[10]) == expected_total
        assert str(out[11]).strip() == expected_summary

def test_summary_sheet_matches_expected(output_wb, source_rows):
    ws = output_wb['Summary']
    headers = [ws.cell(1, c).value for c in range(1, 6)]
    assert headers == SUMMARY_HEADERS
    expected_rows, expected_totals, _ = expected_summary_rows(source_rows)
    actual_rows = []
    grand_total = None
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 6)]
        if all(v is None for v in vals):
            continue
        if str(vals[0]).strip() == 'Grand Total':
            grand_total = (str(vals[0]), str(vals[1]), int(vals[2]), int(vals[3]), int(vals[4]))
        else:
            actual_rows.append((str(vals[0]), str(vals[1]), int(vals[2]), int(vals[3]), int(vals[4])))
    assert actual_rows == expected_rows
    assert grand_total == ('Grand Total', '-', expected_totals[0], expected_totals[1], expected_totals[2])

def test_word_summary_content(source_rows):
    assert Path(OUTPUT_DOCX).exists(), f'Missing output docx: {OUTPUT_DOCX}'
    _, expected_totals, sku_totals = expected_summary_rows(source_rows)
    doc = Document(OUTPUT_DOCX)
    text = '\n'.join(p.text for p in doc.paragraphs if p.text.strip()).lower()
    assert 'price error' in text
    assert 'window error' in text
    for value in expected_totals:
        assert str(value) in text
    top_skus = [sku for sku, total in sorted(sku_totals.items(), key=lambda x: (-x[1], x[0])) if total > 0][:4]
    mentioned = sum(1 for sku in top_skus if sku.lower() in text)
    assert mentioned >= 2, 'Word summary must mention at least two high-priority SKUs'
    assert any(marker in text for marker in ['recommend', 'review', 'should', 'action', 'prioritize'])
