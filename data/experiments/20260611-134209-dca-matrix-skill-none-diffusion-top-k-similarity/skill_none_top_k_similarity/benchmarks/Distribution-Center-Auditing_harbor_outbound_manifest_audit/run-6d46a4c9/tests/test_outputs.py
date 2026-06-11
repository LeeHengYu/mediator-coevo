"""Verification tests for harbor_outbound_manifest_audit."""

from collections import defaultdict
from pathlib import Path
import os

import pytest
from docx import Document
from openpyxl import load_workbook

ROOT_DIR = os.environ.get('TASK_ROOT', '/root')
PLAN_XLSX = f"{ROOT_DIR}/Manifest_Plan.xlsx"
SCAN_XLSX = f"{ROOT_DIR}/Dock_Scan_Log.xlsx"
TEMPLATE_XLSX = f"{ROOT_DIR}/Outbound_Audit_Template.xlsx"
OUTPUT_XLSX = f"{ROOT_DIR}/Outbound_Load_Audit.xlsx"
OUTPUT_DOCX = f"{ROOT_DIR}/Outbound_Load_Brief.docx"

BASE_HEADERS = [
    'Shipment ID',
    'Carton ID',
    'Planned Zone',
    'Route',
    'Expected Weight',
    'Hazmat Flag',
    'Carrier',
    'Wave',
]

FORMATTED_HEADERS = BASE_HEADERS + ['Missing Load Scan', 'Zone Mismatch', 'Total Errors', 'Error Summary']
SUMMARY_HEADERS = ['Route', 'Shipment ID', 'Missing Load Scans', 'Zone Mismatches', 'Total Errors']

@pytest.fixture(scope='module')
def source_data():
    plan_wb = load_workbook(PLAN_XLSX, data_only=True)
    scan_wb = load_workbook(SCAN_XLSX, data_only=True)
    template_wb = load_workbook(TEMPLATE_XLSX, data_only=True)
    plan_ws = plan_wb['PlanLines']
    scan_ws = scan_wb['Scans']
    plan_headers = [plan_ws.cell(1, c).value for c in range(1, 9)]
    assert plan_headers == BASE_HEADERS
    plan_rows = []
    for r in range(2, plan_ws.max_row + 1):
        vals = [plan_ws.cell(r, c).value for c in range(1, 9)]
        if all(v is None for v in vals):
            continue
        plan_rows.append(dict(zip(BASE_HEADERS, vals)))
    latest_loaded = {}
    for r in range(2, scan_ws.max_row + 1):
        shipment_id = scan_ws.cell(r, 1).value
        carton_id = scan_ws.cell(r, 2).value
        scanned_zone = scan_ws.cell(r, 3).value
        scan_ts = scan_ws.cell(r, 4).value
        status = str(scan_ws.cell(r, 5).value or '').strip().upper()
        if shipment_id is None or carton_id is None or status != 'LOADED':
            continue
        key = (str(shipment_id), str(carton_id))
        record = (str(scan_ts), str(scanned_zone))
        if key not in latest_loaded or record[0] > latest_loaded[key][0]:
            latest_loaded[key] = record
    overview_values = [template_wb['Overview'].cell(r, 1).value for r in range(1, template_wb['Overview'].max_row + 1)]
    return plan_rows, latest_loaded, overview_values

@pytest.fixture(scope='module')
def output_wb():
    assert Path(OUTPUT_XLSX).exists(), f'Missing output workbook: {OUTPUT_XLSX}'
    return load_workbook(OUTPUT_XLSX, data_only=True)

def calc_flags(row, latest_loaded):
    key = (str(row['Shipment ID']), str(row['Carton ID']))
    scan = latest_loaded.get(key)
    missing_scan = 1 if scan is None else 0
    zone_mismatch = 0 if scan is None else int(str(scan[1]) != str(row['Planned Zone']))
    total = missing_scan + zone_mismatch
    if total == 0:
        summary = 'None'
    elif missing_scan and zone_mismatch:
        summary = 'Missing Load Scan, Zone Mismatch'
    elif missing_scan:
        summary = 'Missing Load Scan'
    else:
        summary = 'Zone Mismatch'
    return missing_scan, zone_mismatch, total, summary

def expected_summary_rows(plan_rows, latest_loaded):
    agg = defaultdict(lambda: [0, 0, 0])
    shipment_totals = defaultdict(int)
    total_missing = total_zone = total_errors = 0
    for row in plan_rows:
        missing_scan, zone_mismatch, total, _ = calc_flags(row, latest_loaded)
        key = (str(row['Route']), str(row['Shipment ID']))
        agg[key][0] += missing_scan
        agg[key][1] += zone_mismatch
        agg[key][2] += total
        shipment_totals[str(row['Shipment ID'])] += total
        total_missing += missing_scan
        total_zone += zone_mismatch
        total_errors += total
    result_rows = []
    for (route, shipment_id), values in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1])):
        if values[2] > 0:
            result_rows.append((route, shipment_id, values[0], values[1], values[2]))
    return result_rows, (total_missing, total_zone, total_errors), shipment_totals

def test_overview_sheet_preserved(output_wb, source_data):
    _, _, overview_values = source_data
    out_values = [output_wb['Overview'].cell(r, 1).value for r in range(1, output_wb['Overview'].max_row + 1)]
    assert out_values == overview_values

def test_required_sheets_exist(output_wb):
    assert {'Overview', 'RawData', 'Formatted Data', 'Summary'}.issubset(set(output_wb.sheetnames))

def test_rawdata_copies_source_exactly(output_wb, source_data):
    plan_rows, _, _ = source_data
    ws = output_wb['RawData']
    headers = [ws.cell(1, c).value for c in range(1, 9)]
    assert headers == BASE_HEADERS
    actual = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 9)]
        if all(v is None for v in vals):
            continue
        actual.append(vals)
    expected = [[row[h] for h in BASE_HEADERS] for row in plan_rows]
    assert actual == expected

def test_formatted_data_logic(output_wb, source_data):
    plan_rows, latest_loaded, _ = source_data
    ws = output_wb['Formatted Data']
    headers = [ws.cell(1, c).value for c in range(1, 13)]
    assert headers == FORMATTED_HEADERS
    for i, row in enumerate(plan_rows, start=2):
        out = [ws.cell(i, c).value for c in range(1, 13)]
        assert out[:8] == [row[h] for h in BASE_HEADERS]
        expected_missing, expected_zone, expected_total, expected_summary = calc_flags(row, latest_loaded)
        assert int(out[8]) == expected_missing
        assert int(out[9]) == expected_zone
        assert int(out[10]) == expected_total
        assert str(out[11]).strip() == expected_summary

def test_summary_sheet_matches_expected(output_wb, source_data):
    plan_rows, latest_loaded, _ = source_data
    ws = output_wb['Summary']
    headers = [ws.cell(1, c).value for c in range(1, 6)]
    assert headers == SUMMARY_HEADERS
    expected_rows, expected_totals, _ = expected_summary_rows(plan_rows, latest_loaded)
    actual_rows = []
    grand_total = None
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 6)]
        if all(v is None for v in vals):
            continue
        if str(vals[0]).strip() == 'Grand Total':
            grand_total = (str(vals[0]), str(vals[1]), int(vals[2]), int(vals[3]), int(vals[4]))
        else:
            actual_rows.append((str(vals[0]), str(vals[1]), int(vals[2]), int(vals[3]), int(vals[4])))
    assert actual_rows == expected_rows
    assert grand_total == ('Grand Total', '-', expected_totals[0], expected_totals[1], expected_totals[2])

def test_word_summary_content(source_data):
    plan_rows, latest_loaded, _ = source_data
    assert Path(OUTPUT_DOCX).exists(), f'Missing output docx: {OUTPUT_DOCX}'
    _, expected_totals, shipment_totals = expected_summary_rows(plan_rows, latest_loaded)
    doc = Document(OUTPUT_DOCX)
    text = '\n'.join(p.text for p in doc.paragraphs if p.text.strip()).lower()
    assert 'missing load scan' in text
    assert 'zone mismatch' in text
    for value in expected_totals:
        assert str(value) in text
    top_shipments = [sid for sid, total in sorted(shipment_totals.items(), key=lambda x: (-x[1], x[0])) if total > 0][:4]
    mentioned = sum(1 for sid in top_shipments if sid.lower() in text)
    assert mentioned >= 2, 'Word summary must mention at least two high-priority shipment IDs'
    assert any(marker in text for marker in ['recommend', 'review', 'should', 'action', 'prioritize'])
