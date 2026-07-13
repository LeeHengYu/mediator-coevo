from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def clear_range(ws, max_row: int = 80, max_col: int = 20) -> None:
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.value = None


def write_workbook(spec: dict, detail_rows_map: dict, gl_map: dict, output_path: str, template_path: str | None = None) -> None:
    if template_path:
        wb = load_workbook(template_path)
    else:
        wb = Workbook()
        default = wb.active
        wb.remove(default)
        for name in spec["sheet_order"]:
            wb.create_sheet(title=name)

    for name in spec["sheet_order"]:
        if name not in wb.sheetnames:
            wb.create_sheet(title=name)

    if wb.sheetnames != spec["sheet_order"]:
        ordered = []
        for name in spec["sheet_order"]:
            ordered.append(wb[name])
        wb._sheets = ordered

    summary_ws = wb[spec["summary"]["sheet_name"]]
    clear_range(summary_ws)

    detail_infos = []
    for detail in spec["details"]:
        ws = wb[detail["sheet_name"]]
        clear_range(ws)
        detail_infos.append(
            write_detail_sheet(
                ws=ws,
                detail=detail,
                rows=detail_rows_map[detail["sheet_name"]],
                gl_values=gl_map[detail["gl_key"]],
            )
        )

    write_summary_sheet(summary_ws, spec, detail_infos)
    wb.save(output_path)


def write_detail_sheet(ws, detail: dict, rows: list[dict], gl_values: dict) -> dict:
    months = detail["months"]
    term_col_idx = 2 + len(months) * 3 + 1
    notes_col_idx = term_col_idx + 1
    account_col_idx = term_col_idx + 2
    last_end_col_idx = 2 + len(months) * 3

    ws["A1"] = detail["company"]
    ws["A2"] = detail["title"]
    ws["A3"] = detail["subtitle"]

    ws["A4"] = detail["entity_header"]
    ws["B4"] = "Beginning Balance"
    for index, month in enumerate(months):
        base_col = 3 + index * 3
        ws[f"{get_column_letter(base_col)}4"] = month["label"]
        ws[f"{get_column_letter(base_col + 1)}4"] = month["label"]
        ws[f"{get_column_letter(base_col + 2)}4"] = month["label"]
        ws[f"{get_column_letter(base_col)}5"] = month["adds_subheader"]
        ws[f"{get_column_letter(base_col + 1)}5"] = month["release_subheader"]
        ws[f"{get_column_letter(base_col + 2)}5"] = "Ending Balance"
    ws[f"{get_column_letter(term_col_idx)}4"] = detail["term_header"]
    ws[f"{get_column_letter(notes_col_idx)}4"] = detail["notes_header"]
    ws[f"{get_column_letter(account_col_idx)}4"] = detail["account_header"]

    start_row = 6
    current_row = start_row
    for row in rows:
        ws[f"A{current_row}"] = row["entity"]
        ws[f"B{current_row}"] = row["beginning_balance"]
        for index, month in enumerate(months):
            base_col = 3 + index * 3
            slug = month["slug"]
            ws[f"{get_column_letter(base_col)}{current_row}"] = row[f"{slug}_adds"]
            ws[f"{get_column_letter(base_col + 1)}{current_row}"] = row[f"{slug}_release"]
            ws[f"{get_column_letter(base_col + 2)}{current_row}"] = row[f"{slug}_ending_balance"]
        ws[f"{get_column_letter(term_col_idx)}{current_row}"] = row["term_months"]
        ws[f"{get_column_letter(notes_col_idx)}{current_row}"] = row["comments"]
        ws[f"{get_column_letter(account_col_idx)}{current_row}"] = row["account_number"]
        current_row += 1

    end_data_row = current_row - 1
    totals_row = end_data_row + 1
    ending_row = totals_row + 1
    variance_row = ending_row + 1
    gl_row = variance_row + 1

    ws[f"A{totals_row}"] = detail["totals_label"]
    for col_idx in range(2, last_end_col_idx + 1):
        col_letter = get_column_letter(col_idx)
        ws[f"{col_letter}{totals_row}"] = f"=SUM({col_letter}{start_row}:{col_letter}{end_data_row})"
    add_columns = [get_column_letter(3 + index * 3) for index in range(len(months))]
    release_columns = [get_column_letter(4 + index * 3) for index in range(len(months))]
    ws[f"{get_column_letter(term_col_idx)}{totals_row}"] = "=" + "+".join(f"{col}{totals_row}" for col in add_columns)

    ws[f"A{ending_row}"] = "Ending Balance"
    first_end_col = get_column_letter(5)
    ws[f"{first_end_col}{ending_row}"] = f"=B{totals_row}+C{totals_row}-D{totals_row}"
    for index in range(1, len(months)):
        prev_end_col = get_column_letter(5 + (index - 1) * 3)
        add_col = get_column_letter(3 + index * 3)
        release_col = get_column_letter(4 + index * 3)
        end_col = get_column_letter(5 + index * 3)
        ws[f"{end_col}{ending_row}"] = f"={prev_end_col}{totals_row}+{add_col}{totals_row}-{release_col}{totals_row}"
    ws[f"{get_column_letter(term_col_idx)}{ending_row}"] = "=" + "+".join(f"{col}{totals_row}" for col in release_columns)

    last_end_col = get_column_letter(last_end_col_idx)
    term_col = get_column_letter(term_col_idx)
    ws[f"A{variance_row}"] = "Variance"
    ws[f"{term_col}{variance_row}"] = f"={term_col}{gl_row}-{last_end_col}{gl_row}"

    ws[f"A{gl_row}"] = "GL Balance"
    for index, month in enumerate(months):
        end_col = get_column_letter(5 + index * 3)
        ws[f"{end_col}{gl_row}"] = gl_values[month["slug"]]
    ws[f"{term_col}{gl_row}"] = f"={term_col}{totals_row}-{term_col}{ending_row}"

    return {
        "sheet_name": detail["sheet_name"],
        "totals_row": totals_row,
        "ending_row": ending_row,
        "gl_row": gl_row,
        "term_col": term_col,
        "summary_labels": detail["summary_labels"],
    }


def write_summary_sheet(ws, spec: dict, detail_infos: list[dict]) -> None:
    summary = spec["summary"]
    ws["A1"] = summary["company"]
    ws["A2"] = summary["title"]
    ws["A3"] = summary["period_text"]
    ws["A5"] = "Description"
    ws["B5"] = "Amount"

    first, second = detail_infos

    ws["A6"] = first["summary_labels"]["section"]
    ws["A7"] = first["summary_labels"]["total"]
    ws["B7"] = f"='{first['sheet_name']}'!{first['term_col']}{first['totals_row']}"
    ws["A8"] = first["summary_labels"]["ending"]
    ws["B8"] = f"='{first['sheet_name']}'!{first['term_col']}{first['ending_row']}"
    ws["A9"] = first["summary_labels"]["gl"]
    ws["B9"] = f"='{first['sheet_name']}'!{first['term_col']}{first['gl_row']}"

    ws["A11"] = second["summary_labels"]["section"]
    ws["A12"] = second["summary_labels"]["total"]
    ws["B12"] = f"='{second['sheet_name']}'!{second['term_col']}{second['totals_row']}"
    ws["A13"] = second["summary_labels"]["ending"]
    ws["B13"] = f"='{second['sheet_name']}'!{second['term_col']}{second['ending_row']}"
    ws["A14"] = second["summary_labels"]["gl"]
    ws["B14"] = f"='{second['sheet_name']}'!{second['term_col']}{second['gl_row']}"

    ws["A16"] = summary["total_label"]
    ws["B16"] = "=B9+B14"
