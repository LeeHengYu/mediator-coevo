"""Task verifier."""
import os
from typing import Any
from openpyxl import load_workbook


def rows_match(actual_rows, expected_rows, tol=0.01):
    """Compare rows with numeric tolerance for amount fields."""
    if len(actual_rows) != len(expected_rows):
        return False, f"Row count mismatch: {len(actual_rows)} vs {len(expected_rows)}"
    for i, (ar, er) in enumerate(zip(actual_rows, expected_rows)):
        if len(ar) != len(er):
            return False, f"Row {i} column count mismatch: {len(ar)} vs {len(er)}"
        for j, (a, e) in enumerate(zip(ar, er)):
            a_s, e_s = str(a).strip(), str(e).strip()
            # Try numeric comparison
            try:
                a_f, e_f = float(a_s.replace(',', '')), float(e_s.replace(',', ''))
                if abs(a_f - e_f) > tol:
                    return False, f"Row {i} col {j}: numeric mismatch {a_s} vs {e_s}"
            except (ValueError, TypeError):
                if a_s != e_s:
                    return False, f"Row {i} col {j}: text mismatch '{a_s}' vs '{e_s}'"
    return True, ""


def _cell_to_string(value: Any) -> str:
    return '' if value is None else str(value)


def _read_rows(path: str, sheet_name: str):
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[sheet_name]
        rows = []
        for r in range(1, (ws.max_row or 0) + 1):
            rows.append([_cell_to_string(ws.cell(row=r, column=c).value) for c in range(1, (ws.max_column or 0) + 1)])
        return rows
    finally:
        wb.close()


def _sheet_names(path: str):
    wb = load_workbook(path, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def test_outputs():
    output_file = '/app/workspace/fuel_packets.xlsx'
    expected_file = os.path.join(os.path.dirname(__file__), 'fuel_packets_oracle.xlsx')
    assert os.path.exists(output_file), f'Missing output: {output_file}'
    actual_sheets = _sheet_names(output_file)
    assert actual_sheets == ['transactions'], f'Unexpected sheets: {actual_sheets}'
    actual_rows = _read_rows(output_file, 'transactions')
    expected_rows = _read_rows(expected_file, 'transactions')
    assert actual_rows, 'Workbook is empty'
    assert actual_rows[0] == ['batch_name', 'relative_path', 'txn_ref', 'date', 'total_amount'], f'Header mismatch: {actual_rows[0]}'
    data_rows = [row for row in actual_rows[1:] if row and row[0] != '']
    ordering = [row[1] for row in data_rows]
    assert ordering == sorted(ordering), f'Rows not sorted as required: {ordering}'
    ok, msg = rows_match(actual_rows, expected_rows)
    assert ok, f'Workbook differs from oracle: {msg}'
