"""Verification tests for harbor_cycle_count_variance_audit."""

from collections import defaultdict
from pathlib import Path
import os

import pytest
from docx import Document
from openpyxl import load_workbook

ROOT_DIR = os.environ.get('TASK_ROOT', '/root')
PLAN_XLSX = f"{ROOT_DIR}/Cycle_Plan.xlsx"
EVENT_XLSX = f"{ROOT_DIR}/Count_Event_Log.xlsx"
TEMPLATE_XLSX = f"{ROOT_DIR}/Cycle_Template.xlsx"
OUTPUT_XLSX = f"{ROOT_DIR}/Cycle_Count_Variance_Audit.xlsx"
OUTPUT_DOCX = f"{ROOT_DIR}/Cycle_Count_Variance_Brief.docx"

BASE_HEADERS = [
    'Facility',
    'Session ID',
    'Bin ID',
    'Product ID',
    'Expected Qty',
    'Allowed Variance',
    'Approval Needed',
]

FORMATTED_HEADERS = BASE_HEADERS + ['Missing Final Count', 'Approval Gap', 'Total Errors', 'Error Summary']
SUMMARY_HEADERS = ['Facility', 'Session ID', 'Missing Final Counts', 'Approval Gaps', 'Total Errors']


@pytest.fixture(scope='module')
def source_data():
    plan_wb = load_workbook(PLAN_XLSX, data_only=True)
    event_wb = load_workbook(EVENT_XLSX, data_only=True)
    template_wb = load_workbook(TEMPLATE_XLSX, data_only=True)

    plan_ws = plan_wb['PlanLines']
    event_ws = event_wb['Events']

    plan_headers = [plan_ws.cell(1, c).value for c in range(1, 8)]
    assert plan_headers == BASE_HEADERS
    plan_rows = []
    for r in range(2, plan_ws.max_row + 1):
        vals = [plan_ws.cell(r, c).value for c in range(1, 8)]
        if all(v is None for v in vals):
            continue
        plan_rows.append(dict(zip(BASE_HEADERS, vals)))

    latest_final = {}
    for r in range(2, event_ws.max_row + 1):
        facility = event_ws.cell(r, 1).value
        session = event_ws.cell(r, 2).value
        bin_id = event_ws.cell(r, 3).value
        event_time = event_ws.cell(r, 4).value
        event_type = str(event_ws.cell(r, 5).value or '').strip().upper()
        count_qty = event_ws.cell(r, 6).value
        if facility is None or session is None or bin_id is None or count_qty is None:
            continue
        if event_type != 'FINAL':
            continue
        key = (str(facility), str(session), str(bin_id))
        record = (str(event_time), float(count_qty))
        if key not in latest_final or record[0] > latest_final[key][0]:
            latest_final[key] = record

    template_overview = []
    template_ws = template_wb['Overview']
    for r in range(1, template_ws.max_row + 1):
        row = [template_ws.cell(r, c).value for c in range(1, template_ws.max_column + 1)]
        template_overview.append(row)

    return plan_rows, latest_final, template_overview


@pytest.fixture(scope='module')
def output_wb():
    assert Path(OUTPUT_XLSX).exists(), f'Missing output workbook: {OUTPUT_XLSX}'
    return load_workbook(OUTPUT_XLSX, data_only=True)


def calc_flags(row, latest_final):
    key = (str(row['Facility']), str(row['Session ID']), str(row['Bin ID']))
    event = latest_final.get(key)
    missing = 1 if event is None else 0
    approval_needed = str(row['Approval Needed'] or '').strip().upper() == 'YES'
    gap = 0
    if event is not None and approval_needed:
        expected = float(row['Expected Qty'])
        actual = event[1]
        allowed = float(row['Allowed Variance'])
        if abs(expected - actual) > allowed:
            gap = 1
    total = missing + gap
    if total == 0:
        summary = 'None'
    elif missing and gap:
        summary = 'Missing Final Count, Approval Gap'
    elif missing:
        summary = 'Missing Final Count'
    else:
        summary = 'Approval Gap'
    return missing, gap, total, summary


def expected_summary_rows(plan_rows, latest_final):
    agg = defaultdict(lambda: [0, 0, 0])
    session_totals = defaultdict(int)
    total_missing = total_gap = total_errors = 0
    for row in plan_rows:
        missing, gap, total, _ = calc_flags(row, latest_final)
        key = (str(row['Facility']), str(row['Session ID']))
        agg[key][0] += missing
        agg[key][1] += gap
        agg[key][2] += total
        session_totals[key] += total
        total_missing += missing
        total_gap += gap
        total_errors += total
    result_rows = []
    for (facility, session), values in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1])):
        if values[2] > 0:
            result_rows.append((facility, session, values[0], values[1], values[2]))
    return result_rows, (total_missing, total_gap, total_errors), session_totals


def test_overview_preserved(output_wb, source_data):
    _, _, template_overview = source_data
    assert 'Overview' in output_wb.sheetnames
    ws = output_wb['Overview']
    actual = []
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        actual.append(row)
    assert actual == template_overview


def test_required_sheets_exist(output_wb):
    assert {'Overview', 'RawData', 'Formatted Data', 'Summary'}.issubset(set(output_wb.sheetnames))


def test_rawdata_copies_source_exactly(output_wb, source_data):
    plan_rows, _, _ = source_data
    ws = output_wb['RawData']
    headers = [ws.cell(1, c).value for c in range(1, 8)]
    assert headers == BASE_HEADERS
    actual = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 8)]
        if all(v is None for v in vals):
            continue
        actual.append(vals)
    expected = [[row[h] for h in BASE_HEADERS] for row in plan_rows]
    assert actual == expected


def test_formatted_data_logic(output_wb, source_data):
    plan_rows, latest_final, _ = source_data
    ws = output_wb['Formatted Data']
    headers = [ws.cell(1, c).value for c in range(1, 12)]
    assert headers == FORMATTED_HEADERS
    for i, row in enumerate(plan_rows, start=2):
        out = [ws.cell(i, c).value for c in range(1, 12)]
        assert out[:7] == [row[h] for h in BASE_HEADERS]
        expected_missing, expected_gap, expected_total, expected_summary = calc_flags(row, latest_final)
        assert int(out[7]) == expected_missing
        assert int(out[8]) == expected_gap
        assert int(out[9]) == expected_total
        assert str(out[10]).strip() == expected_summary


def test_summary_sheet_matches_expected(output_wb, source_data):
    plan_rows, latest_final, _ = source_data
    ws = output_wb['Summary']
    headers = [ws.cell(1, c).value for c in range(1, 6)]
    assert headers == SUMMARY_HEADERS
    expected_rows, expected_totals, _ = expected_summary_rows(plan_rows, latest_final)
    actual_rows = []
    grand_total = None
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 6)]
        if all(v is None for v in vals):
            continue
        if str(vals[0]).strip() == 'Grand Total':
            grand_total = (str(vals[0]), str(vals[1]), int(vals[2]), int(vals[3]), int(vals[4]))
        else:
            actual_rows.append((str(vals[0]), str(vals[1]), int(vals[2]), int(vals[3]), int(vals[4])))
    assert actual_rows == expected_rows
    assert grand_total == ('Grand Total', '-', expected_totals[0], expected_totals[1], expected_totals[2])


def test_word_summary_content(source_data):
    plan_rows, latest_final, _ = source_data
    assert Path(OUTPUT_DOCX).exists(), f'Missing output docx: {OUTPUT_DOCX}'
    _, expected_totals, session_totals = expected_summary_rows(plan_rows, latest_final)
    doc = Document(OUTPUT_DOCX)
    text = '\n'.join(p.text for p in doc.paragraphs if p.text.strip()).lower()
    assert 'missing final count' in text
    assert 'approval gap' in text
    for value in expected_totals:
        assert str(value) in text
    top_sessions = [f"{fac}-{sess}" for (fac, sess), total in sorted(session_totals.items(), key=lambda x: (-x[1], x[0])) if total > 0][:4]
    mentioned = sum(1 for s in top_sessions if s.lower() in text or s.replace('-', ' ').lower() in text)
    assert mentioned >= 2, 'Word summary must mention at least two high-priority facility-session combinations'
    assert any(marker in text for marker in ['recommend', 'review', 'should', 'action', 'prioritize'])
