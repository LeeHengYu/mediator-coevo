"""Verification tests for harbor_timesheet_policy_audit."""

from collections import defaultdict
from pathlib import Path
import os

import pytest
from docx import Document
from openpyxl import load_workbook

ROOT_DIR = os.environ.get('TASK_ROOT', '/root')
SOURCE_XLSX = f"{ROOT_DIR}/Timesheet_Submissions.xlsx"
OUTPUT_XLSX = f"{ROOT_DIR}/Timesheet_Compliance_Audit.xlsx"
OUTPUT_DOCX = f"{ROOT_DIR}/Timesheet_Compliance_Brief.docx"

BASE_HEADERS = [
    'Week Ending',
    'Employee ID',
    'Role',
    'Hours Worked',
    'Break Minutes',
    'Approval Code',
    'Project Code',
    'Manager',
]

FORMATTED_HEADERS = BASE_HEADERS + ['Break Deficit', 'Approval Missing', 'Total Errors', 'Error Summary']
SUMMARY_HEADERS = ['Employee ID', 'Week Ending', 'Break Deficits', 'Approval Gaps', 'Total Errors']

@pytest.fixture(scope='module')
def source_and_rules():
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    entries_ws = wb['Entries']
    rules_ws = wb['BreakRules']
    entry_headers = [entries_ws.cell(1, c).value for c in range(1, 9)]
    assert entry_headers == BASE_HEADERS
    rows = []
    for r in range(2, entries_ws.max_row + 1):
        vals = [entries_ws.cell(r, c).value for c in range(1, 9)]
        if all(v is None for v in vals):
            continue
        rows.append(dict(zip(BASE_HEADERS, vals)))
    rules = {}
    for r in range(2, rules_ws.max_row + 1):
        role = rules_ws.cell(r, 1).value
        if role is None:
            continue
        rules[str(role)] = {
            'min_break': float(rules_ws.cell(r, 2).value),
            'overtime_threshold': float(rules_ws.cell(r, 3).value),
        }
    assert rows and rules
    return rows, rules

@pytest.fixture(scope='module')
def output_wb():
    assert Path(OUTPUT_XLSX).exists(), f'Missing output workbook: {OUTPUT_XLSX}'
    return load_workbook(OUTPUT_XLSX, data_only=True)

def calc_flags(row, rules):
    rule = rules[str(row['Role'])]
    break_deficit = 1 if float(row['Break Minutes']) < rule['min_break'] else 0
    approval_code = str(row['Approval Code'] or '').strip()
    approval_missing = 1 if float(row['Hours Worked']) > rule['overtime_threshold'] and not approval_code else 0
    total = break_deficit + approval_missing
    if total == 0:
        summary = 'None'
    elif break_deficit and approval_missing:
        summary = 'Break Deficit, Approval Missing'
    elif break_deficit:
        summary = 'Break Deficit'
    else:
        summary = 'Approval Missing'
    return break_deficit, approval_missing, total, summary

def expected_summary_rows(rows, rules):
    agg = defaultdict(lambda: [0, 0, 0])
    employee_totals = defaultdict(int)
    total_break = total_approval = total_errors = 0
    for row in rows:
        break_deficit, approval_missing, total, _ = calc_flags(row, rules)
        key = (str(row['Employee ID']), str(row['Week Ending']))
        agg[key][0] += break_deficit
        agg[key][1] += approval_missing
        agg[key][2] += total
        employee_totals[str(row['Employee ID'])] += total
        total_break += break_deficit
        total_approval += approval_missing
        total_errors += total
    result_rows = []
    for (employee_id, week_ending), values in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1])):
        if values[2] > 0:
            result_rows.append((employee_id, week_ending, values[0], values[1], values[2]))
    return result_rows, (total_break, total_approval, total_errors), employee_totals

def test_required_sheets_exist(output_wb):
    assert {'RawData', 'Formatted Data', 'Summary'}.issubset(set(output_wb.sheetnames))

def test_rawdata_copies_source_exactly(output_wb, source_and_rules):
    rows, _ = source_and_rules
    ws = output_wb['RawData']
    headers = [ws.cell(1, c).value for c in range(1, 9)]
    assert headers == BASE_HEADERS
    actual = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 9)]
        if all(v is None for v in vals):
            continue
        actual.append(vals)
    expected = [[row[h] for h in BASE_HEADERS] for row in rows]
    assert actual == expected

def test_formatted_data_logic(output_wb, source_and_rules):
    rows, rules = source_and_rules
    ws = output_wb['Formatted Data']
    headers = [ws.cell(1, c).value for c in range(1, 13)]
    assert headers == FORMATTED_HEADERS
    for i, row in enumerate(rows, start=2):
        out = [ws.cell(i, c).value for c in range(1, 13)]
        assert out[:8] == [row[h] for h in BASE_HEADERS]
        expected_break, expected_approval, expected_total, expected_summary = calc_flags(row, rules)
        assert int(out[8]) == expected_break
        assert int(out[9]) == expected_approval
        assert int(out[10]) == expected_total
        assert str(out[11]).strip() == expected_summary

def test_summary_sheet_matches_expected(output_wb, source_and_rules):
    rows, rules = source_and_rules
    ws = output_wb['Summary']
    headers = [ws.cell(1, c).value for c in range(1, 6)]
    assert headers == SUMMARY_HEADERS
    expected_rows, expected_totals, _ = expected_summary_rows(rows, rules)
    actual_rows = []
    grand_total = None
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 6)]
        if all(v is None for v in vals):
            continue
        if str(vals[0]).strip() == 'Grand Total':
            grand_total = (str(vals[0]), str(vals[1]), int(vals[2]), int(vals[3]), int(vals[4]))
        else:
            actual_rows.append((str(vals[0]), str(vals[1]), int(vals[2]), int(vals[3]), int(vals[4])))
    assert actual_rows == expected_rows
    assert grand_total == ('Grand Total', '-', expected_totals[0], expected_totals[1], expected_totals[2])

def test_word_summary_content(source_and_rules):
    rows, rules = source_and_rules
    assert Path(OUTPUT_DOCX).exists(), f'Missing output docx: {OUTPUT_DOCX}'
    _, expected_totals, employee_totals = expected_summary_rows(rows, rules)
    doc = Document(OUTPUT_DOCX)
    text = '\n'.join(p.text for p in doc.paragraphs if p.text.strip()).lower()
    assert 'break deficit' in text
    assert 'approval missing' in text
    for value in expected_totals:
        assert str(value) in text
    top_employees = [emp for emp, total in sorted(employee_totals.items(), key=lambda x: (-x[1], x[0])) if total > 0][:4]
    mentioned = sum(1 for emp in top_employees if emp.lower() in text)
    assert mentioned >= 2, 'Word summary must mention at least two high-priority employee IDs'
    assert any(marker in text for marker in ['recommend', 'review', 'should', 'action', 'prioritize'])
