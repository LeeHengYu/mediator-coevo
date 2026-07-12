"""Verification tests for harbor_returns_disposition_audit."""

from collections import defaultdict
from pathlib import Path
import os

import pytest
from docx import Document
from openpyxl import load_workbook

ROOT_DIR = os.environ.get('TASK_ROOT', '/root')
PLAN_XLSX = f"{ROOT_DIR}/Return_Plan.xlsx"
EVENT_XLSX = f"{ROOT_DIR}/Disposition_Event_Log.xlsx"
ALIAS_XLSX = f"{ROOT_DIR}/Disposition_Alias.xlsx"
OUTPUT_XLSX = f"{ROOT_DIR}/Returns_Disposition_Audit.xlsx"
OUTPUT_DOCX = f"{ROOT_DIR}/Returns_Disposition_Brief.docx"

BASE_HEADERS = [
    'Return ID',
    'Line ID',
    'Planned Disposition',
    'Reason Code',
    'Requested Qty',
    'Warehouse',
    'Carrier',
    'Lane',
]

FORMATTED_HEADERS = BASE_HEADERS + ['Missing Final Event', 'Disposition Mismatch', 'Total Errors', 'Error Summary']
SUMMARY_HEADERS = ['Warehouse', 'Carrier', 'Missing Final Events', 'Disposition Mismatches', 'Total Errors']


@pytest.fixture(scope='module')
def source_data():
    plan_wb = load_workbook(PLAN_XLSX, data_only=True)
    event_wb = load_workbook(EVENT_XLSX, data_only=True)
    alias_wb = load_workbook(ALIAS_XLSX, data_only=True)
    plan_ws = plan_wb['PlanLines']
    event_ws = event_wb['Events']
    alias_ws = alias_wb['AliasMap']
    plan_headers = [plan_ws.cell(1, c).value for c in range(1, 9)]
    assert plan_headers == BASE_HEADERS
    plan_rows = []
    for r in range(2, plan_ws.max_row + 1):
        vals = [plan_ws.cell(r, c).value for c in range(1, 9)]
        if all(v is None for v in vals):
            continue
        plan_rows.append(dict(zip(BASE_HEADERS, vals)))
    latest_completed = {}
    for r in range(2, event_ws.max_row + 1):
        return_id = event_ws.cell(r, 1).value
        line_id = event_ws.cell(r, 2).value
        event_time = event_ws.cell(r, 3).value
        status = str(event_ws.cell(r, 5).value or '').strip().upper()
        final_disp = event_ws.cell(r, 6).value
        if return_id is None or line_id is None or status != 'COMPLETED':
            continue
        key = (str(return_id), str(line_id))
        record = (str(event_time), str(final_disp))
        if key not in latest_completed or record[0] > latest_completed[key][0]:
            latest_completed[key] = record
    alias_map = {}
    for r in range(2, alias_ws.max_row + 1):
        alias = alias_ws.cell(r, 1).value
        standard = alias_ws.cell(r, 2).value
        if alias is None or standard is None:
            continue
        alias_map[str(alias).strip().upper()] = str(standard).strip().upper()
    return plan_rows, latest_completed, alias_map


@pytest.fixture(scope='module')
def output_wb():
    assert Path(OUTPUT_XLSX).exists(), f'Missing output workbook: {OUTPUT_XLSX}'
    return load_workbook(OUTPUT_XLSX, data_only=True)


def normalize_disposition(raw, alias_map):
    key = str(raw or '').strip().upper()
    return alias_map.get(key, key)


def calc_flags(row, latest_completed, alias_map):
    key = (str(row['Return ID']), str(row['Line ID']))
    event = latest_completed.get(key)
    missing = 1 if event is None else 0
    planned = str(row['Planned Disposition'] or '').strip().upper()
    mismatch = 0
    if event is not None:
        normalized = normalize_disposition(event[1], alias_map)
        mismatch = 1 if normalized != planned else 0
    total = missing + mismatch
    if total == 0:
        summary = 'None'
    elif missing and mismatch:
        summary = 'Missing Final Event, Disposition Mismatch'
    elif missing:
        summary = 'Missing Final Event'
    else:
        summary = 'Disposition Mismatch'
    return missing, mismatch, total, summary


def expected_summary_rows(plan_rows, latest_completed, alias_map):
    agg = defaultdict(lambda: [0, 0, 0])
    return_totals = defaultdict(int)
    total_missing = total_mismatch = total_errors = 0
    for row in plan_rows:
        missing, mismatch, total, _ = calc_flags(row, latest_completed, alias_map)
        key = (str(row['Warehouse']), str(row['Carrier']))
        agg[key][0] += missing
        agg[key][1] += mismatch
        agg[key][2] += total
        return_totals[str(row['Return ID'])] += total
        total_missing += missing
        total_mismatch += mismatch
        total_errors += total
    result_rows = []
    for (warehouse, carrier), values in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1])):
        if values[2] > 0:
            result_rows.append((warehouse, carrier, values[0], values[1], values[2]))
    return result_rows, (total_missing, total_mismatch, total_errors), return_totals


def test_required_sheets_exist(output_wb):
    assert {'RawData', 'Formatted Data', 'Summary'}.issubset(set(output_wb.sheetnames))


def test_rawdata_copies_source_exactly(output_wb, source_data):
    plan_rows, _, _ = source_data
    ws = output_wb['RawData']
    headers = [ws.cell(1, c).value for c in range(1, 9)]
    assert headers == BASE_HEADERS
    actual = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 9)]
        if all(v is None for v in vals):
            continue
        actual.append(vals)
    expected = [[row[h] for h in BASE_HEADERS] for row in plan_rows]
    assert actual == expected


def test_formatted_data_logic(output_wb, source_data):
    plan_rows, latest_completed, alias_map = source_data
    ws = output_wb['Formatted Data']
    headers = [ws.cell(1, c).value for c in range(1, 13)]
    assert headers == FORMATTED_HEADERS
    for i, row in enumerate(plan_rows, start=2):
        out = [ws.cell(i, c).value for c in range(1, 13)]
        assert out[:8] == [row[h] for h in BASE_HEADERS]
        expected_missing, expected_mismatch, expected_total, expected_summary = calc_flags(row, latest_completed, alias_map)
        assert int(out[8]) == expected_missing
        assert int(out[9]) == expected_mismatch
        assert int(out[10]) == expected_total
        assert str(out[11]).strip() == expected_summary


def test_summary_sheet_matches_expected(output_wb, source_data):
    plan_rows, latest_completed, alias_map = source_data
    ws = output_wb['Summary']
    headers = [ws.cell(1, c).value for c in range(1, 6)]
    assert headers == SUMMARY_HEADERS
    expected_rows, expected_totals, _ = expected_summary_rows(plan_rows, latest_completed, alias_map)
    actual_rows = []
    grand_total = None
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 6)]
        if all(v is None for v in vals):
            continue
        if str(vals[0]).strip() == 'Grand Total':
            grand_total = (str(vals[0]), str(vals[1]), int(vals[2]), int(vals[3]), int(vals[4]))
        else:
            actual_rows.append((str(vals[0]), str(vals[1]), int(vals[2]), int(vals[3]), int(vals[4])))
    assert actual_rows == expected_rows
    assert grand_total == ('Grand Total', '-', expected_totals[0], expected_totals[1], expected_totals[2])


def test_word_summary_content(source_data):
    plan_rows, latest_completed, alias_map = source_data
    assert Path(OUTPUT_DOCX).exists(), f'Missing output docx: {OUTPUT_DOCX}'
    _, expected_totals, return_totals = expected_summary_rows(plan_rows, latest_completed, alias_map)
    doc = Document(OUTPUT_DOCX)
    text = '\n'.join(p.text for p in doc.paragraphs if p.text.strip()).lower()
    assert 'missing final event' in text
    assert 'disposition mismatch' in text
    for value in expected_totals:
        assert str(value) in text
    top_returns = [rid for rid, total in sorted(return_totals.items(), key=lambda x: (-x[1], x[0])) if total > 0][:4]
    mentioned = sum(1 for rid in top_returns if rid.lower() in text)
    assert mentioned >= 2, 'Word summary must mention at least two high-priority return IDs'
    assert any(marker in text for marker in ['recommend', 'review', 'should', 'action', 'prioritize'])
