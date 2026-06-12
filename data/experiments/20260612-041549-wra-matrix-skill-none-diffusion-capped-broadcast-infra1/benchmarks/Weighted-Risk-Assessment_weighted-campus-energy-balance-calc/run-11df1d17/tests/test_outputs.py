import csv
import glob
import subprocess
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook

EXCEL_FILE = Path("/root/output/result.xlsx")
CSV_PATTERN = "/root/output/sheet.csv.*"
COLUMNS = ["H", "I", "J", "K", "L"]
TOLERANCE = 0.02
YELLOW_RGBS = {"00FFF2CC", "FFF2CC", "00FFF2CC"}

_recalculated_path = None


def _has_cached_values(path: Path) -> bool:
    """Check if the workbook has cached values for formula cells."""
    wb = load_workbook(path, data_only=True)
    ws = wb["Task"]
    for ref in ["H12", "H35", "H50"]:
        val = ws[ref].value
        if not isinstance(val, (int, float)):
            wb.close()
            return False
    wb.close()
    return True


def _recalculate_with_ssconvert(src: Path) -> Path:
    """Use ssconvert to recalculate formulas and write to a temp file."""
    tmp = Path(tempfile.mkdtemp()) / "recalculated.xlsx"
    try:
        subprocess.run(
            ["ssconvert", str(src), str(tmp), "--recalculate"],
            capture_output=True,
            check=True,
        )
    except subprocess.CalledProcessError as exc:
        stderr = (exc.stderr or b"").decode(errors="ignore")
        if "Unknown option --recalculate" not in stderr:
            raise
        subprocess.run(["ssconvert", str(src), str(tmp)], capture_output=True, check=True)
    return tmp


def _get_reliable_workbook_path() -> Path:
    """Return path to workbook with cached values, recalculating if needed."""
    global _recalculated_path
    if _recalculated_path is not None and _recalculated_path.exists():
        return _recalculated_path
    if _has_cached_values(EXCEL_FILE):
        return EXCEL_FILE
    _recalculated_path = _recalculate_with_ssconvert(EXCEL_FILE)
    return _recalculated_path

EXPECTED_TOP = {"12": [188.0, 194.8, 201.5, 208.4, 215.2], "13": [222.6, 228.1, 233.2, 238.0, 242.4], "14": [175.3, 180.7, 186.4, 192.3, 198.6], "15": [260.0, 267.4, 275.2, 283.3, 291.8], "16": [205.8, 212.6, 219.1, 225.4, 231.8], "17": [191.5, 197.0, 202.8, 208.9, 215.1]}
EXPECTED_MID = {"19": [195.2, 200.0, 204.5, 208.8, 212.8], "20": [211.0, 216.5, 222.1, 227.8, 233.8], "21": [179.5, 184.0, 188.8, 193.9, 199.1], "22": [239.0, 245.2, 251.1, 256.8, 262.1], "23": [198.2, 204.4, 210.9, 217.6, 224.7], "24": [196.5, 202.0, 207.8, 214.0, 220.5]}
EXPECTED_BASE = {"26": [310.0, 319.0, 328.0, 337.0, 346.0], "27": [355.0, 364.0, 373.0, 382.0, 391.0], "28": [290.0, 299.0, 308.0, 317.0, 326.0], "29": [420.0, 431.0, 442.0, 453.0, 464.0], "30": [340.0, 349.0, 358.0, 367.0, 376.0], "31": [305.0, 314.0, 323.0, 332.0, 341.0]}
EXPECTED_NET = {"35": [-2.3225806451612865, -1.6300940438871439, -0.9146341463414633, -0.11869436201780584, 0.6936416184971033], "36": [3.2676056338028157, 3.1868131868131853, 2.9758713136729207, 2.6701570680628244, 2.1994884910485917], "37": [-1.4482758620689615, -1.10367892976589, -0.779220779220781, -0.5047318611987364, -0.15337423312883436], "38": [5.0, 5.150812064965195, 5.452488687782804, 5.849889624724062, 6.400862068965515], "39": [2.2352941176470655, 2.349570200573063, 2.290502793296086, 2.125340599455044, 1.8882978723404318], "40": [-1.639344262295082, -1.5923566878980893, -1.5479876160990713, -1.5361445783132512, -1.5835777126099724]}
EXPECTED_STATS = {"42": [-2.3225806451612865, -1.6300940438871439, -1.5479876160990713, -1.5361445783132512, -1.5835777126099724], "43": [5.0, 5.150812064965195, 5.452488687782804, 5.849889624724062, 6.400862068965515], "44": [0.393509127789052, 0.6229456354035865, 0.7556410070376526, 1.003323118718619, 1.2909697454187676], "45": [0.8487831636540918, 1.0601776318000533, 1.2461700421817492, 1.414302748452023, 1.574223017518806], "46": [-1.591577162238552, -1.4701872483650393, -0.8807808045612928, -0.4082224864035037, 0.05837972977765005], "47": [3.0095277547638783, 2.9775024402531547, 2.804529183578712, 2.533952950910879, 2.1216908363715516]}
EXPECTED_WEIGHTED = [1.1782178217821802, 1.3728323699421954, 1.5478424015009367, 1.7093235831809874, 1.8672014260249545]

_csv_cache = None


def workbook(data_only=True):
    path = _get_reliable_workbook_path() if data_only else EXCEL_FILE
    return load_workbook(path, data_only=data_only)


def task_sheet(wb):
    return wb["Task"]


def find_task_csv():
    files = sorted(glob.glob(CSV_PATTERN))
    if not files:
        return None
    wb = workbook(data_only=False)
    idx = wb.sheetnames.index("Task")
    wb.close()
    candidate = f"/root/output/sheet.csv.{idx}"
    return candidate if Path(candidate).exists() else files[0]


def load_csv():
    global _csv_cache
    if _csv_cache is not None:
        return _csv_cache
    _csv_cache = {}
    csv_file = find_task_csv()
    if csv_file is None:
        return _csv_cache
    with open(csv_file, encoding="utf-8", errors="ignore") as handle:
        reader = csv.reader(handle)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                ref = f"{chr(ord('A') + col_idx - 1)}{row_idx}"
                if value in (None, ""):
                    _csv_cache[ref] = None
                    continue
                try:
                    _csv_cache[ref] = float(value)
                except ValueError:
                    _csv_cache[ref] = value
    return _csv_cache


def value(ws, ref):
    direct = ws[ref].value
    if isinstance(direct, (int, float)):
        return float(direct)
    cached = load_csv().get(ref)
    if isinstance(cached, (int, float)):
        return float(cached)
    return direct


def assert_matrix(ws, expected_map, label):
    errors = []
    for row, expected_values in expected_map.items():
        for idx, col in enumerate(COLUMNS):
            ref = f"{col}{row}"
            actual = value(ws, ref)
            expected = expected_values[idx]
            if not isinstance(actual, (int, float)) or abs(actual - expected) > TOLERANCE:
                errors.append(f"{ref}: expected {expected}, got {actual}")
    assert not errors, f"{label} mismatches:\n" + "\n".join(errors)


def test_file_and_sheet_structure():
    assert EXCEL_FILE.exists(), f"Missing output workbook: {EXCEL_FILE}"
    wb = workbook()
    assert wb.sheetnames == ["Task", "Data"]
    ws = task_sheet(wb)
    assert ws["A1"].value is not None
    wb.close()


def test_lookup_blocks():
    wb = workbook()
    ws = task_sheet(wb)
    assert_matrix(ws, EXPECTED_TOP, "top block")
    assert_matrix(ws, EXPECTED_MID, "middle block")
    assert_matrix(ws, EXPECTED_BASE, "base block")
    wb.close()


def test_derived_values_and_stats():
    wb = workbook()
    ws = task_sheet(wb)
    assert_matrix(ws, EXPECTED_NET, "net metric")
    assert_matrix(ws, EXPECTED_STATS, "statistics")
    weighted_errors = []
    for idx, col in enumerate(COLUMNS):
        ref = f"{col}50"
        actual = value(ws, ref)
        expected = EXPECTED_WEIGHTED[idx]
        if not isinstance(actual, (int, float)) or abs(actual - expected) > TOLERANCE:
            weighted_errors.append(f"{ref}: expected {expected}, got {actual}")
    wb.close()
    assert not weighted_errors, "weighted mean mismatches:\n" + "\n".join(weighted_errors)


def test_formulas_present_in_editable_ranges():
    wb = workbook(data_only=False)
    ws = task_sheet(wb)
    missing = []
    for row in list(range(12, 18)) + list(range(19, 25)) + list(range(26, 32)) + list(range(35, 41)) + [42, 43, 44, 45, 46, 47, 50]:
        for col in COLUMNS:
            ref = f"{col}{row}"
            cell_value = ws[ref].value
            if not (isinstance(cell_value, str) and cell_value.startswith("=")):
                missing.append(f"{ref}: {cell_value}")
    for col in COLUMNS:
        formula = ws[f"{col}50"].value
        if "SUMPRODUCT" not in str(formula).upper():
            missing.append(f"{col}50 missing SUMPRODUCT: {formula}")
    wb.close()
    assert not missing, "Missing formulas:\n" + "\n".join(missing)


def test_template_formatting_preserved():
    wb = workbook(data_only=False)
    ws = task_sheet(wb)
    for ref in ["H12", "L31", "H35", "L47", "H50"]:
        fill = ws[ref].fill
        rgb = getattr(fill.fgColor, "rgb", None)
        assert fill.patternType == "solid", f"{ref} lost yellow fill"
        assert rgb in YELLOW_RGBS, f"{ref} fill changed: {rgb}"
    wb.close()


def test_no_excel_errors_or_macros():
    errors = []
    csv_values = load_csv()
    for ref, value_ in csv_values.items():
        if isinstance(value_, str) and any(token in value_ for token in ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]):
            errors.append(f"{ref}: {value_}")
    with zipfile.ZipFile(EXCEL_FILE, "r") as archive:
        macros = [name for name in archive.namelist() if "vbaProject" in name or name.endswith(".bin")]
    if macros:
        errors.append(f"Macros found: {macros}")
    assert not errors, "Validation errors:\n" + "\n".join(errors)
