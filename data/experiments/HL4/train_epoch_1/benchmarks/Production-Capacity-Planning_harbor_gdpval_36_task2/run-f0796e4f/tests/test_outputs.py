import re
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook

PLAN_FILE = Path("/root/mill_catch_up_plan.xlsx")
SUMMARY_FILE = Path("/root/mill_catch_up_summary.txt")
REFERENCE_FILE = Path("/root/mill_demand_sheet.xlsx")

PERIOD_START = 1
PERIOD_END = 52
INITIAL_PERIOD1_TOTAL = 538.08
RATE_PER_DAY = 25.0


def _clean_label(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def load_reference_demands() -> Dict[int, float]:
    wb = load_workbook(REFERENCE_FILE, data_only=True)
    ws = wb["Mill"] if "Mill" in wb.sheetnames else wb.active

    out: Dict[int, float] = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2, values_only=True):
        if not row:
            continue
        period_val = row[0]
        demand_val = row[1] if len(row) > 1 else None
        if period_val is None or demand_val is None:
            continue
        try:
            period = int(round(float(period_val)))
            demand = float(demand_val)
        except (TypeError, ValueError):
            continue
        out[period] = demand

    for period in range(PERIOD_START, PERIOD_END + 1):
        assert period in out, f"Reference demand missing for period {period}"
    return out


def load_plan_rows() -> List[Tuple[int, int, float, float, float, float, float]]:
    assert PLAN_FILE.exists(), f"Output workbook missing: {PLAN_FILE}"
    wb = load_workbook(PLAN_FILE, data_only=True)
    ws = wb["Plan"] if "Plan" in wb.sheetnames else wb.active

    expected_headers = [
        "Period",
        "Days Worked",
        "Scheduled Demand (Std Hrs)",
        "Weekly Capacity (Std Hrs)",
        "Start of Period Past Due (Std Hrs)",
        "End of Period Backlog/Buffer (Std Hrs)",
        "Overtime Hours",
    ]
    actual_headers = [ws.cell(1, i).value for i in range(1, 8)]
    assert actual_headers == expected_headers, (
        f"Header mismatch.\nExpected: {expected_headers}\nActual:   {actual_headers}"
    )

    rows: List[Tuple[int, int, float, float, float, float, float]] = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 8)]
        if all(v is None for v in vals):
            continue
        period = int(round(float(vals[0])))
        days = int(round(float(vals[1])))
        demand = float(vals[2])
        capacity = float(vals[3])
        start_past_due = float(vals[4])
        end_backlog_buffer = float(vals[5])
        overtime = float(vals[6])
        rows.append((period, days, demand, capacity, start_past_due, end_backlog_buffer, overtime))

    assert len(rows) == 52, f"Expected 52 period rows (periods 1..52), found {len(rows)}"
    return rows


def expected_days(calc_start: float, demand: float) -> int:
    start_past_due = max(0.0, calc_start)
    if start_past_due > 0.01:
        if calc_start + demand - (RATE_PER_DAY * 5) <= 0.0:
            return 5
        return 6
    return 4 if demand <= 125.0 else 5


def test_plan_math_and_policy():
    ref_demands = load_reference_demands()
    rows = load_plan_rows()

    expected_periods = list(range(PERIOD_START, PERIOD_END + 1))
    actual_periods = [r[0] for r in rows]
    assert actual_periods == expected_periods, "Periods must be continuous from 1 to 52 with no gaps"

    calc_start = INITIAL_PERIOD1_TOTAL - ref_demands[PERIOD_START]
    prev_end = None
    for idx, (period, days, demand, capacity, start_past_due, end_backlog_buffer, overtime) in enumerate(rows):
        assert days in {4, 5, 6}, f"Period {period}: days must be one of 4/5/6"

        # Demand must match source data.
        assert abs(demand - ref_demands[period]) <= 0.01, (
            f"Period {period}: demand {demand} does not match reference {ref_demands[period]}"
        )

        # Period 1 initial condition.
        if period == PERIOD_START:
            assert abs(start_past_due + demand - INITIAL_PERIOD1_TOTAL) <= 0.01, (
                "Period 1 must satisfy Start of Period Past Due + Scheduled Demand = 538.08"
            )

        # Carryover consistency.
        if prev_end is not None:
            assert abs(calc_start - prev_end) <= 0.01, f"Period {period}: calc carryover mismatch"

        # Start of period past due should be max(0, signed start).
        assert abs(start_past_due - max(0.0, calc_start)) <= 0.01, (
            f"Period {period}: Start of Period Past Due should equal max(0, prior end)"
        )

        # Capacity and backlog equations.
        assert abs(capacity - RATE_PER_DAY * days) <= 0.01, f"Period {period}: capacity formula mismatch"
        expected_end = calc_start + demand - capacity
        assert abs(end_backlog_buffer - expected_end) <= 0.01, f"Period {period}: end backlog formula mismatch"

        # Overtime formula.
        expected_ot = 10.0 * max(0, days - 4)
        assert abs(overtime - expected_ot) <= 0.01, f"Period {period}: overtime formula mismatch"

        # Deterministic day-selection policy.
        exp_days = expected_days(calc_start, demand)
        assert days == exp_days, f"Period {period}: expected {exp_days} day(s), got {days}"

        prev_end = end_backlog_buffer
        calc_start = end_backlog_buffer


def test_summary_file():
    assert SUMMARY_FILE.exists(), f"Summary file missing: {SUMMARY_FILE}"
    rows = load_plan_rows()

    first_period_5 = next((period for period, days, *_ in rows if days == 5), None)
    first_period_4 = next((period for period, days, *_ in rows if days == 4), None)

    content = SUMMARY_FILE.read_text(encoding="utf-8").strip()
    lines = content.splitlines()
    assert len(lines) == 3, "Summary file must contain exactly 3 lines"

    m5 = re.fullmatch(r"First_Week_5_Days:\s*(\d+|N/A)", lines[0].strip())
    m4 = re.fullmatch(r"First_Week_4_Days:\s*(\d+|N/A)", lines[1].strip())
    assert m5, "Line 1 must match: First_Week_5_Days: <period-number-or-N/A>"
    assert m4, "Line 2 must match: First_Week_4_Days: <period-number-or-N/A>"

    got_5 = m5.group(1)
    got_4 = m4.group(1)
    exp_5 = str(first_period_5) if first_period_5 is not None else "N/A"
    exp_4 = str(first_period_4) if first_period_4 is not None else "N/A"
    assert got_5 == exp_5, f"First_Week_5_Days mismatch: expected {exp_5}, got {got_5}"
    assert got_4 == exp_4, f"First_Week_4_Days mismatch: expected {exp_4}, got {got_4}"

    assert lines[2].startswith("Summary: "), "Line 3 must start with 'Summary: '"
    summary_text = lines[2][len("Summary: "):].strip()
    assert summary_text, "Summary text cannot be empty"

    words = re.findall(r"\S+", summary_text)
    assert len(words) <= 60, f"Summary exceeds 60 words ({len(words)})"
    sentence_count = len(re.findall(r"[.!?]", summary_text))
    assert sentence_count <= 3, f"Summary exceeds 3 sentences ({sentence_count})"

    # Must mention both step-down period numbers or N/A strings.
    assert got_5 in summary_text, "Summary must mention the 5-day step-down period"
    assert got_4 in summary_text, "Summary must mention the 4-day step-down period"
