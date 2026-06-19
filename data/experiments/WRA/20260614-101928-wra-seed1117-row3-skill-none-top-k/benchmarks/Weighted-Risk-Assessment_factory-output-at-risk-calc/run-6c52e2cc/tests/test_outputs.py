import csv
import glob
import subprocess
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook

EXCEL_FILE = Path("/root/output/result.xlsx")
CSV_PATTERN = "/root/output/sheet.csv.*"
COLUMNS = ["H", "I", "J", "K", "L"]
TOLERANCE = 0.02
YELLOW_RGBS = {"00FFF2CC", "FFF2CC", "00FFF2CC"}

_recalculated_path = None


def _has_cached_values(path: Path) -> bool:
    """Check if the workbook has cached values for formula cells."""
    wb = load_workbook(path, data_only=True)
    ws = wb["Task"]
    for ref in ["H12", "H35", "H50"]:
        val = ws[ref].value
        if not isinstance(val, (int, float)):
            wb.close()
            return False
    wb.close()
    return True


def _recalculate_with_ssconvert(src: Path) -> Path:
    """Use ssconvert to recalculate formulas and write to a temp file."""
    tmp = Path(tempfile.mkdtemp()) / "recalculated.xlsx"
    try:
        subprocess.run(
            ["ssconvert", str(src), str(tmp), "--recalculate"],
            capture_output=True,
            check=True,
        )
    except subprocess.CalledProcessError as exc:
        stderr = (exc.stderr or b"").decode(errors="ignore")
        if "Unknown option --recalculate" not in stderr:
            raise
        subprocess.run(["ssconvert", str(src), str(tmp)], capture_output=True, check=True)
    return tmp


def _get_reliable_workbook_path() -> Path:
    """Return path to workbook with cached values, recalculating if needed."""
    global _recalculated_path
    if _recalculated_path is not None and _recalculated_path.exists():
        return _recalculated_path
    if _has_cached_values(EXCEL_FILE):
        return EXCEL_FILE
    _recalculated_path = _recalculate_with_ssconvert(EXCEL_FILE)
    return _recalculated_path

EXPECTED_TOP = {"12": [705.0, 718.5, 732.2, 745.8, 759.4], "13": [612.2, 623.0, 633.5, 643.8, 653.9], "14": [540.0, 550.5, 561.2, 572.0, 582.9], "15": [660.8, 673.2, 685.6, 698.1, 710.8], "16": [498.6, 507.9, 517.4, 527.0, 536.8], "17": [580.3, 591.5, 602.9, 614.4, 626.0]}
EXPECTED_MID = {"19": [676.6, 695.3, 714.8, 734.7, 755.1], "20": [616.6, 625.8, 634.8, 643.5, 652.0], "21": [552.3, 560.9, 569.4, 577.8, 586.1], "22": [618.0, 631.8, 646.1, 660.8, 675.7], "23": [486.2, 496.2, 506.5, 517.0, 527.8], "24": [591.9, 601.2, 610.4, 619.7, 629.0]}
EXPECTED_BASE = {"26": [820.0, 842.0, 864.0, 886.0, 908.0], "27": [690.0, 708.0, 726.0, 744.0, 762.0], "28": [615.0, 631.0, 647.0, 663.0, 679.0], "29": [780.0, 800.0, 820.0, 840.0, 860.0], "30": [560.0, 575.0, 590.0, 605.0, 620.0], "31": [640.0, 657.0, 674.0, 691.0, 708.0]}
EXPECTED_NET = {"35": [3.463414634146339, 2.755344418052262, 2.0138888888888995, 1.2528216704288835, 0.47356828193832096], "36": [-0.6376811594202866, -0.39548022598869415, -0.17906336088153646, 0.04032258064515518, 0.2493438320209944], "37": [-1.9999999999999927, -1.648177496038031, -1.2673879443585676, -0.8748114630467503, -0.47128129602357083], "38": [5.487179487179481, 5.175000000000011, 4.817073170731708, 4.440476190476199, 4.081395348837199], "39": [2.21428571428572, 2.03478260869565, 1.8474576271186403, 1.6528925619834711, 1.4516129032258065], "40": [-1.8125000000000038, -1.476407914764086, -1.112759643916914, -0.7670043415340185, -0.423728813559322]}
EXPECTED_STATS = {"42": [-1.9999999999999927, -1.648177496038031, -1.2673879443585676, -0.8748114630467503, -0.47128129602357083], "43": [5.487179487179481, 5.175000000000011, 4.817073170731708, 4.440476190476199, 4.081395348837199], "44": [0.7883022774327169, 0.8196511913534781, 0.8341971331185519, 0.6465721255370194, 0.36145605697965766], "45": [1.119116446031876, 1.0741768983261855, 1.0198681229303717, 0.9574495331588233, 0.8934850427399046], "46": [-1.5187952898550745, -1.206175992570238, -0.8793355731580696, -0.5651726109892251, -0.2554606521642429], "47": [3.1511324041811846, 2.575203965713109, 1.9722810734463347, 1.5528748390948242, 1.2071017479039352]}
EXPECTED_WEIGHTED = [1.347137637028015, 1.2675053406123937, 1.1756537838463361, 1.074734703093247, 0.9720079347586464]

_csv_cache = None


def workbook(data_only=True):
    path = _get_reliable_workbook_path() if data_only else EXCEL_FILE
    return load_workbook(path, data_only=data_only)


def task_sheet(wb):
    return wb["Task"]


def find_task_csv():
    files = sorted(glob.glob(CSV_PATTERN))
    if not files:
        return None
    wb = workbook(data_only=False)
    idx = wb.sheetnames.index("Task")
    wb.close()
    candidate = f"/root/output/sheet.csv.{idx}"
    return candidate if Path(candidate).exists() else files[0]


def load_csv():
    global _csv_cache
    if _csv_cache is not None:
        return _csv_cache
    _csv_cache = {}
    csv_file = find_task_csv()
    if csv_file is None:
        return _csv_cache
    with open(csv_file, encoding="utf-8", errors="ignore") as handle:
        reader = csv.reader(handle)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                ref = f"{chr(ord('A') + col_idx - 1)}{row_idx}"
                if value in (None, ""):
                    _csv_cache[ref] = None
                    continue
                try:
                    _csv_cache[ref] = float(value)
                except ValueError:
                    _csv_cache[ref] = value
    return _csv_cache


def value(ws, ref):
    direct = ws[ref].value
    if isinstance(direct, (int, float)):
        return float(direct)
    cached = load_csv().get(ref)
    if isinstance(cached, (int, float)):
        return float(cached)
    return direct


def assert_matrix(ws, expected_map, label):
    errors = []
    for row, expected_values in expected_map.items():
        for idx, col in enumerate(COLUMNS):
            ref = f"{col}{row}"
            actual = value(ws, ref)
            expected = expected_values[idx]
            if not isinstance(actual, (int, float)) or abs(actual - expected) > TOLERANCE:
                errors.append(f"{ref}: expected {expected}, got {actual}")
    assert not errors, f"{label} mismatches:\n" + "\n".join(errors)


def test_file_and_sheet_structure():
    assert EXCEL_FILE.exists(), f"Missing output workbook: {EXCEL_FILE}"
    wb = workbook()
    assert wb.sheetnames == ["Task", "Data"]
    ws = task_sheet(wb)
    assert ws["A1"].value is not None
    wb.close()


def test_lookup_blocks():
    wb = workbook()
    ws = task_sheet(wb)
    assert_matrix(ws, EXPECTED_TOP, "top block")
    assert_matrix(ws, EXPECTED_MID, "middle block")
    assert_matrix(ws, EXPECTED_BASE, "base block")
    wb.close()


def test_derived_values_and_stats():
    wb = workbook()
    ws = task_sheet(wb)
    assert_matrix(ws, EXPECTED_NET, "net metric")
    assert_matrix(ws, EXPECTED_STATS, "statistics")
    weighted_errors = []
    for idx, col in enumerate(COLUMNS):
        ref = f"{col}50"
        actual = value(ws, ref)
        expected = EXPECTED_WEIGHTED[idx]
        if not isinstance(actual, (int, float)) or abs(actual - expected) > TOLERANCE:
            weighted_errors.append(f"{ref}: expected {expected}, got {actual}")
    wb.close()
    assert not weighted_errors, "weighted mean mismatches:\n" + "\n".join(weighted_errors)


def test_formulas_present_in_editable_ranges():
    wb = workbook(data_only=False)
    ws = task_sheet(wb)
    missing = []
    for row in list(range(12, 18)) + list(range(19, 25)) + list(range(26, 32)) + list(range(35, 41)) + [42, 43, 44, 45, 46, 47, 50]:
        for col in COLUMNS:
            ref = f"{col}{row}"
            cell_value = ws[ref].value
            if not (isinstance(cell_value, str) and cell_value.startswith("=")):
                missing.append(f"{ref}: {cell_value}")
    for col in COLUMNS:
        formula = ws[f"{col}50"].value
        if "SUMPRODUCT" not in str(formula).upper():
            missing.append(f"{col}50 missing SUMPRODUCT: {formula}")
    wb.close()
    assert not missing, "Missing formulas:\n" + "\n".join(missing)


def test_template_formatting_preserved():
    wb = workbook(data_only=False)
    ws = task_sheet(wb)
    for ref in ["H12", "L31", "H35", "L47", "H50"]:
        fill = ws[ref].fill
        rgb = getattr(fill.fgColor, "rgb", None)
        assert fill.patternType == "solid", f"{ref} lost yellow fill"
        assert rgb in YELLOW_RGBS, f"{ref} fill changed: {rgb}"
    wb.close()


def test_no_excel_errors_or_macros():
    errors = []
    csv_values = load_csv()
    for ref, value_ in csv_values.items():
        if isinstance(value_, str) and any(token in value_ for token in ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]):
            errors.append(f"{ref}: {value_}")
    with zipfile.ZipFile(EXCEL_FILE, "r") as archive:
        macros = [name for name in archive.namelist() if "vbaProject" in name or name.endswith(".bin")]
    if macros:
        errors.append(f"Macros found: {macros}")
    assert not errors, "Validation errors:\n" + "\n".join(errors)
