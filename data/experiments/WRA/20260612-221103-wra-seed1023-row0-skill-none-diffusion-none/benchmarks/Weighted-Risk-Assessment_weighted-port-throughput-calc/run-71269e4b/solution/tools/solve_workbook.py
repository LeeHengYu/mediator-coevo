from __future__ import annotations

import shutil
import statistics
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _percentile(data: list[float], pct: float) -> float:
    """Excel PERCENTILE (exclusive) equivalent using linear interpolation."""
    s = sorted(data)
    n = len(s)
    if n == 1:
        return s[0]
    k = (n - 1) * pct / 100.0
    f = int(k)
    c = k - f
    if f + 1 < n:
        return s[f] + c * (s[f + 1] - s[f])
    return s[f]


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("Usage: solve_workbook.py <input.xlsx> <output.xlsx>")

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(input_path, output_path)

    wb = load_workbook(output_path)
    task = wb["Task"]
    data_ws = wb["Data"]

    COLUMNS = ["H", "I", "J", "K", "L"]
    COL_IDX = list(range(8, 13))

    # ── Read Data sheet for lookups ─────────────────────────────────────
    # Data rows 21..38: D=lookup key, H..L=values
    data_map: dict[str, list[float]] = {}
    for r in range(21, 39):
        key = data_ws[f"D{r}"].value
        if key is None:
            continue
        key = str(key)
        vals = []
        for c in COLUMNS:
            v = data_ws[f"{c}{r}"].value
            vals.append(float(v) if isinstance(v, (int, float)) else 0.0)
        data_map[key] = vals

    # Task row 10: scenario headers for column matching
    task_headers = [task[f"{c}10"].value for c in COLUMNS]
    # Data row 4: scenario headers for column matching
    data_headers = [data_ws[f"{c}4"].value for c in COLUMNS]

    # Build column index mapping: task header -> position in data values
    col_mapping: dict[str, int] = {}  # task_header -> index in data vals
    for i, th in enumerate(task_headers):
        for j, dh in enumerate(data_headers):
            if th == dh:
                col_mapping[th] = j
                break

    # ── Compute all values in Python ────────────────────────────────────
    computed: dict[str, float] = {}
    lookup_rows = list(range(12, 18)) + list(range(19, 25)) + list(range(26, 32))

    for row in lookup_rows:
        key = str(task[f"D{row}"].value or "")
        src = data_map.get(key, [0.0] * 5)
        for i, col in enumerate(COLUMNS):
            th = task_headers[i]
            idx = col_mapping.get(th, i)
            computed[f"{col}{row}"] = src[idx]

    # Derived rows 35-40
    for offset in range(6):
        top_row = 12 + offset
        mid_row = 19 + offset
        base_row = 26 + offset
        row = 35 + offset
        for col in COLUMNS:
            t = computed[f"{col}{top_row}"]
            m = computed[f"{col}{mid_row}"]
            b = computed[f"{col}{base_row}"]
            computed[f"{col}{row}"] = (t - m) / b * 100 if b != 0 else 0.0

    # Statistics rows 42-47, 50
    for col in COLUMNS:
        vals = [computed[f"{col}{r}"] for r in range(35, 41)]
        computed[f"{col}42"] = min(vals)
        computed[f"{col}43"] = max(vals)
        computed[f"{col}44"] = statistics.median(vals)
        computed[f"{col}45"] = statistics.mean(vals)
        computed[f"{col}46"] = _percentile(vals, 25)
        computed[f"{col}47"] = _percentile(vals, 75)
        weights = [computed[f"{col}{r}"] for r in range(26, 32)]
        w_sum = sum(weights)
        if w_sum != 0:
            computed[f"{col}50"] = sum(v * w for v, w in zip(vals, weights)) / w_sum
        else:
            computed[f"{col}50"] = 0.0

    # ── Write formulas to cells ─────────────────────────────────────────
    for row in lookup_rows:
        for col_idx in COL_IDX:
            col = get_column_letter(col_idx)
            task[f"{col}{row}"] = (
                f"=INDEX(Data!$H$21:$L$38,"
                f"MATCH($D{row},Data!$D$21:$D$38,0),"
                f"MATCH({col}$10,Data!$H$4:$L$4,0))"
            )

    for offset, row in enumerate(range(35, 41)):
        top_row = 12 + offset
        mid_row = 19 + offset
        base_row = 26 + offset
        for col_idx in COL_IDX:
            col = get_column_letter(col_idx)
            task[f"{col}{row}"] = f"=({col}{top_row}-{col}{mid_row})/{col}{base_row}*100"

    for col_idx in COL_IDX:
        col = get_column_letter(col_idx)
        value_range = f"{col}35:{col}40"
        weight_range = f"{col}26:{col}31"
        task[f"{col}42"] = f"=MIN({value_range})"
        task[f"{col}43"] = f"=MAX({value_range})"
        task[f"{col}44"] = f"=MEDIAN({value_range})"
        task[f"{col}45"] = f"=AVERAGE({value_range})"
        task[f"{col}46"] = f"=PERCENTILE({value_range},0.25)"
        task[f"{col}47"] = f"=PERCENTILE({value_range},0.75)"
        task[f"{col}50"] = f"=SUMPRODUCT({value_range},{weight_range})/SUM({weight_range})"

    wb.save(output_path)
    wb.close()

    # ── Inject cached values into xlsx XML ──────────────────────────────
    _inject_values(output_path, computed)


def _inject_values(xlsx_path: Path, computed: dict[str, float]) -> None:
    """Add <v> cached values to formula cells so openpyxl data_only=True reads them."""
    NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_pfx = f"{{{NS}}}"

    # Find which worksheet file corresponds to "Task" sheet
    sheet_xml_name = None
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        for name in zf.namelist():
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                with zf.open(name) as f:
                    content = f.read().decode("utf-8")
                    # The first sheet is typically Task; check for formula patterns
                    if "MATCH" in content:
                        sheet_xml_name = name
                        break

    if sheet_xml_name is None:
        # Fallback: assume Task is sheet1
        sheet_xml_name = "xl/worksheets/sheet1.xml"

    tmp_path = xlsx_path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == sheet_xml_name:
                    raw = zin.read(item.filename)
                    tree = ET.fromstring(raw)
                    for row_elem in tree.iter(f"{ns_pfx}row"):
                        for cell_elem in row_elem.iter(f"{ns_pfx}c"):
                            ref = cell_elem.get("r", "")
                            if ref in computed:
                                # Remove existing <v> if any
                                for v_elem in cell_elem.findall(f"{ns_pfx}v"):
                                    cell_elem.remove(v_elem)
                                # Add new <v>
                                v_elem = ET.SubElement(cell_elem, f"{ns_pfx}v")
                                v_elem.text = str(computed[ref])
                    new_xml = ET.tostring(tree, encoding="unicode", xml_declaration=True)
                    zout.writestr(item, new_xml.encode("utf-8"))
                else:
                    zout.writestr(item, zin.read(item.filename))

    shutil.move(str(tmp_path), str(xlsx_path))


if __name__ == "__main__":
    main()
